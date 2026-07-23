"""
UV台帐管理系统 - Flask Backend
提供API: 文件上传、台帐校准、历史查询、数据筛选
"""
import os
import sys
import json
import uuid
import traceback
import webbrowser
import threading
from datetime import datetime

from flask import (
    Flask, request, jsonify, send_file, render_template, make_response
)

from calibrator import (
    calibrate, generate_calibrated_ledger, generate_calibration_report,
    normalize_teaching_point, load_enrollment_orders
)
from db import (
    save_calibration_run, get_all_runs, get_run_by_id,
    get_latest_run, get_daily_stats, get_filtered_students,
    get_distinct_values, init_db, is_continue_fall,
    get_period_teaching_point_matrix, get_class_schedule,
    get_teacher_detail, get_teacher_list, get_student_detail, get_db,
    set_data_dir
)

# ── 数据目录：打包模式使用 ~/Library/Application Support/UV Dashboard/ ──
def _get_data_dir():
    """获取数据目录路径。打包为 .app 时使用系统 Application Support 目录"""
    if getattr(sys, 'frozen', False):
        # PyInstaller 打包模式
        base = os.path.expanduser('~/Library/Application Support/UV Dashboard')
    else:
        # 开发模式
        base = os.path.dirname(os.path.abspath(__file__))
    os.makedirs(base, exist_ok=True)
    return base

DATA_DIR = _get_data_dir()
set_data_dir(DATA_DIR)

app = Flask(__name__,
            template_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
            if not getattr(sys, 'frozen', False)
            else os.path.join(sys._MEIPASS, 'templates'),
            static_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
            if not getattr(sys, 'frozen', False)
            else None)

app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB
app.config['UPLOAD_FOLDER'] = os.path.join(DATA_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(DATA_DIR, 'outputs')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

init_db()

# ═══ 期次时间表（开课日期 / 结课日期）═══
# 零期: 6/29开课, 7/4-5休息, 7/10结课
# 一期: 7/13开课, 7/18-19休息, 7/24结课
# 二期: 7/27开课, 8/7结课
# 三期: 8/10开课, 8/21结课
PERIOD_SCHEDULE = {
    '零期': {'start': (6, 29), 'end': (7, 10)},
    '一期': {'start': (7, 13), 'end': (7, 24)},
    '二期': {'start': (7, 27), 'end': (8, 7)},
    '三期': {'start': (8, 10), 'end': (8, 21)},
}

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def get_period_start_str(period_name, year=None):
    """获取期次开课日期字符串 (YYYY-MM-DD)"""
    if year is None:
        year = datetime.now().year
    m, d = PERIOD_SCHEDULE[period_name]['start']
    return f'{year}-{m:02d}-{d:02d}'


def get_baseline_run_id(conn, period_name, current_run_id, year=None):
    """
    获取某期次开课前最近一次快照的 run_id（排除当前 run）。
    如果找不到开课前快照（或只有当前 run），返回 None（空基线），
    意味着该期所有续费都是"当期转化"。
    """
    p_start = get_period_start_str(period_name, year)
    pre_r = conn.execute(
        'SELECT run_id FROM calibration_runs WHERE date(run_time) < ? AND run_id != ? ORDER BY run_time DESC LIMIT 1',
        (p_start, current_run_id)
    ).fetchone()
    return pre_r['run_id'] if pre_r else None


def _classify_renewal_by_pay_time(fall_pay_time_str, period_name, year=None):
    """
    根据秋季续费订单的支付时间判断是"开课前续费"还是"当期转化"。
    - 支付时间 < 期次开课日期 → 'pre' (开课前续费)
    - 支付时间 >= 期次开课日期 → 'new' (当期转化)
    - 无支付时间 → 'pre' (保守归为开课前续费，如从原台帐继承的续秋标记)
    Returns: 'pre' or 'new'
    """
    if not fall_pay_time_str:
        return 'pre'
    try:
        pay_time = datetime.fromisoformat(str(fall_pay_time_str))
    except (ValueError, TypeError):
        return 'pre'
    if year is None:
        year = datetime.now().year
    schedule = PERIOD_SCHEDULE.get(period_name)
    if not schedule:
        return 'pre'
    m, d = schedule['start']
    period_start = datetime(year, m, d)
    return 'pre' if pay_time < period_start else 'new'

PERIOD_ORDER = ['零期', '一期', '二期', '三期']


def get_period_order():
    return PERIOD_ORDER


def get_auto_period():
    now = datetime.now()
    t_md = now.month * 100 + now.day
    if t_md >= 810: return '三期'
    elif t_md >= 727: return '二期'
    elif t_md >= 713: return '一期'
    else: return '零期'


ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    if not filename or '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS


def get_file_ext(filename):
    """安全提取文件扩展名（小写）"""
    if not filename or '.' not in filename:
        return ''
    return filename.rsplit('.', 1)[1].lower()


def json_response(data, status=200):
    resp = make_response(json.dumps(data, ensure_ascii=False, default=str), status)
    resp.headers['Content-Type'] = 'application/json; charset=utf-8'
    return resp


def _compute_filtered_new_count(conn, rid, point_filter, subj_filter, period_filter, teacher_filter):
    """
    按筛选条件计算某次校准 run 中的新增学员 PV 数。

    新增学员来源：change_records 中 change_type='新增学员' 的 uid + subject 对。
    PV 逻辑：每个新增学员的每个在读科目满足筛选条件时计 1（不去重 uid）。
    """
    import json as j

    # 1. 从 change_records 取出新增学员的 uid + subject 对
    where_sql = 'run_id = ? AND change_type = ?'
    params = [rid, '新增学员']
    if subj_filter:
        where_sql += ' AND subject IN (%s)' % ','.join('?' * len(subj_filter))
        params.extend(subj_filter)

    rows = conn.execute(
        f'SELECT DISTINCT uid, subject FROM change_records WHERE {where_sql}',
        params
    ).fetchall()
    if not rows:
        return 0

    # 构建 uid -> set(subjects) 映射
    uid_subjects = {}
    for r in rows:
        uid_subjects.setdefault(r['uid'], set()).add(r['subject'])

    # 2. 取这些 uid 的 snapshot，按教学点/期次/主讲筛选，计 PV
    uids = list(uid_subjects.keys())
    placeholders = ','.join('?' * len(uids))
    snap_rows = conn.execute(
        f'SELECT uid, teaching_point, subjects_json FROM student_snapshots WHERE run_id = ? AND uid IN ({placeholders})',
        (rid,) + tuple(uids)
    ).fetchall()

    pv_count = 0
    for s in snap_rows:
        pt = (s['teaching_point'] or '').strip()
        if point_filter and pt not in point_filter:
            continue
        try:
            subs = j.loads(s['subjects_json'] or '{}')
        except Exception:
            continue
        new_subjects = uid_subjects.get(s['uid'], set())
        for subj, sd in subs.items():
            if not isinstance(sd, dict):
                continue
            # 只统计该学员新增的科目
            if subj not in new_subjects:
                continue
            if subj_filter and subj not in subj_filter:
                continue
            period = (sd.get('period') or '').strip()
            if period_filter:
                plist = [p.strip() for p in period.split('/') if p.strip()]
                if not any(p in period_filter for p in plist):
                    continue
            teacher = (sd.get('teacher') or '').strip()
            if teacher_filter and teacher not in teacher_filter:
                continue
            status = (sd.get('status') or '').strip()
            if '在读' in status:
                pv_count += 1
    return pv_count


# ── 页面 ──

@app.route('/')
def index():
    return render_template('index.html')


# ── API: 台帐校准 ──

@app.route('/api/calibrate', methods=['POST'])
def api_calibrate():
    """上传招生明细+原台帐，执行校准"""
    try:
        enrollment_file = request.files.get('enrollment')
        ledger_file = request.files.get('ledger')
        fall_file = request.files.get('fall_orders')

        if not enrollment_file or not ledger_file:
            return json_response({'error': '请上传招生明细和原UV台帐两个文件'}, 400)

        if not allowed_file(enrollment_file.filename) or not allowed_file(ledger_file.filename):
            return json_response({'error': '仅支持 .xlsx 文件'}, 400)

        # 保存上传文件（使用随机安全文件名，保留原扩展名）
        run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        enf_ext = get_file_ext(enrollment_file.filename)
        led_ext = get_file_ext(ledger_file.filename)
        enf_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{run_id}_enrollment.{enf_ext}')
        led_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{run_id}_ledger.{led_ext}')
        enrollment_file.save(enf_path)
        ledger_file.save(led_path)

        fall_path = None
        if fall_file and allowed_file(fall_file.filename):
            fall_ext = get_file_ext(fall_file.filename)
            fall_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{run_id}_fall.{fall_ext}')
            fall_file.save(fall_path)

        # 执行校准
        result = calibrate(enf_path, led_path, fall_path)

        # 生成输出文件
        ledger_out = os.path.join(app.config['OUTPUT_FOLDER'], f'{run_id}_UV台帳校准版.xlsx')
        report_out = os.path.join(app.config['OUTPUT_FOLDER'], f'{run_id}_台帳校准报告.xlsx')

        generate_calibrated_ledger(result, led_path, ledger_out)
        generate_calibration_report(result, report_out)

        # 保存到数据库
        save_calibration_run(run_id, result['stats'], result,
                            ledger_out, report_out,
                            enrollment_file.filename, ledger_file.filename)

        # 学员变动总览（按学员维度）
        student_changes = result.get('student_changes', [])
        # 按类型分组供前端展示
        new_student_list = [sc for sc in student_changes if sc['change_type'] == '新增学员']
        changed_list = [sc for sc in student_changes if sc['change_type'] == '信息变更']

        return json_response({
            'success': True,
            'run_id': run_id,
            'stats': result['stats'],
            'student_changes': student_changes,
            'new_students_list': new_student_list,
            'changed_list': changed_list,
            'change_details': result.get('change_details', []),
            'anomalies': result.get('anomalies', []),
            'files': {
                'ledger': f'/api/download/{run_id}/ledger',
                'report': f'/api/download/{run_id}/report',
            }
        })

    except Exception as e:
        traceback.print_exc()
        return json_response({'error': f'校准失败: {str(e)}'}, 500)


@app.route('/api/download/<run_id>/<file_type>')
def api_download(run_id, file_type):
    """下载校准输出文件"""
    if file_type == 'ledger':
        path = os.path.join(app.config['OUTPUT_FOLDER'], f'{run_id}_UV台帳校准版.xlsx')
        filename = f'UV台帳_校准版_{run_id}.xlsx'
    elif file_type == 'report':
        path = os.path.join(app.config['OUTPUT_FOLDER'], f'{run_id}_台帳校准报告.xlsx')
        filename = f'台帳校准报告_{run_id}.xlsx'
    else:
        return json_response({'error': '未知文件类型'}, 400)

    if not os.path.exists(path):
        return json_response({'error': '文件不存在'}, 404)

    return send_file(path, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── API: 校准监控 ──

@app.route('/api/runs', methods=['GET'])
def api_get_runs():
    """获取所有校准运行记录"""
    runs = get_all_runs()
    # Parse stats_json
    for r in runs:
        if r.get('stats_json'):
            try:
                r['stats'] = json.loads(r['stats_json'])
            except:
                r['stats'] = {}
    return json_response({'runs': runs})


@app.route('/api/runs/<run_id>', methods=['GET'])
def api_get_run_detail(run_id):
    """获取单次运行详情"""
    data = get_run_by_id(run_id)
    if data is None:
        return json_response({'error': '记录不存在'}, 404)
    return json_response(data)


@app.route('/api/runs/<run_id>/changes', methods=['GET'])
def api_get_run_changes(run_id):
    """获取某次校准的变动记录详情"""
    conn = get_db()
    changes = conn.execute(
        'SELECT * FROM change_records WHERE run_id = ? ORDER BY id', (run_id,)
    ).fetchall()
    conn.close()
    return json_response({'changes': [dict(c) for c in changes]})


@app.route('/api/daily-stats', methods=['GET'])
def api_daily_stats():
    """获取每日统计趋势"""
    stats = get_daily_stats()
    return json_response({'daily': stats})


# ── API: 数据总览 ──

@app.route('/api/overview/filters', methods=['GET'])
def api_overview_filters():
    """获取筛选器可选项"""
    latest = get_latest_run()
    if latest is None:
        return json_response({'filters': {}})

    run_id = latest['run']['run_id']
    filters = {
        'teaching_points': get_distinct_values(run_id, 'teaching_point'),
        'periods': get_distinct_values(run_id, 'period'),
        'teachers': get_distinct_values(run_id, 'teacher'),
        'advisors': get_distinct_values(run_id, 'advisor'),
    }
    return json_response({'filters': filters})


@app.route('/api/overview/students', methods=['POST'])
def api_overview_students():
    """多维度筛选学员列表"""
    try:
        latest = get_latest_run()
        if latest is None:
            return json_response({'students': [], 'total': 0})

        run_id = latest['run']['run_id']
        filters = request.get_json() or {}

        # Parse subjects_json for each student
        students = get_filtered_students(run_id, filters)
        import json as j
        for s in students:
            try:
                s['subjects'] = j.loads(s.get('subjects_json', '{}'))
                s['manual'] = j.loads(s.get('manual_json', '{}'))
            except:
                s['subjects'] = {}
                s['manual'] = {}

        return json_response({
            'students': students,
            'total': len(students),
            'run_time': latest['run']['run_time'],
        })
    except Exception as e:
        traceback.print_exc()
        return json_response({'error': str(e)}, 500)

@app.route('/api/trends/data', methods=['GET'])
def api_trends_data():
    """业务趋势数据 - 基于历史校准快照，支持多维筛选"""
    import json as j
    from datetime import datetime, timedelta
    conn = get_db()

    def _is_cf(v):
        if not v: return False
        v = str(v).strip()
        return v in ('是', '已续', 'T', 'AC', 'AL', 'true', '1', '√')

    # 读取筛选参数
    start_date    = request.args.get('start_date', '')
    end_date      = request.args.get('end_date', '')
    subjects_param = request.args.get('subjects', '')
    points_param   = request.args.get('teaching_points', '')
    periods_param  = request.args.get('periods', '')
    teachers_param = request.args.get('teachers', '')

    subj_filter   = [x.strip() for x in subjects_param.split(',') if x.strip()] if subjects_param else []
    point_filter  = [x.strip() for x in points_param.split(',') if x.strip()] if points_param else []
    period_filter = [x.strip() for x in periods_param.split(',') if x.strip()] if periods_param else []
    teacher_filter = [x.strip() for x in teachers_param.split(',') if x.strip()] if teachers_param else []

    # 构建基础查询（默认近7天）
    sql = 'SELECT * FROM calibration_runs WHERE 1=1'
    params = []
    if start_date:
        sql += ' AND date(run_time) >= ?'
        params.append(start_date)
    if end_date:
        sql += ' AND date(run_time) <= ?'
        params.append(end_date)
    if not start_date and not end_date:
        default_start = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        sql += ' AND date(run_time) >= ?'
        params.append(default_start)
    sql += ' ORDER BY run_time ASC'

    runs = conn.execute(sql, params).fetchall()
    ALL_SUBJS = ['雪球思维', '悦读创作', '双语素养']

    # 按天聚合：同一天多次校准取最后一次
    day_map = {}  # {date: last_run_row}
    for row in runs:
        run = dict(row)
        day = run['run_time'][:10]  # YYYY-MM-DD
        if day not in day_map or run['run_time'] > day_map[day]['run_time']:
            day_map[day] = run

    trends = []
    for day in sorted(day_map.keys()):
        run = day_map[day]
        rid = run['run_id']
        rtime = day  # 按天展示
        snapshots = conn.execute(
            'SELECT * FROM student_snapshots WHERE run_id = ?', (rid,)
        ).fetchall()

        by_subject  = {s: {'active':0,'refund':0,'renewal':0} for s in ALL_SUBJS}
        by_point    = {}
        by_period   = {}
        by_teacher = {}

        for s_row in snapshots:
            s = dict(s_row)
            pt = (s.get('teaching_point') or '').strip()
            if point_filter and pt not in point_filter:
                continue
            try:
                subs = j.loads(s.get('subjects_json') or '{}')
            except:
                subs = {}

            for subj, sd in subs.items():
                if not isinstance(sd, dict): continue
                if subj_filter and subj not in subj_filter: continue

                period  = (sd.get('period') or '').strip()
                teacher = (sd.get('teacher') or '').strip()

                # 期次筛选（跨期用/拼接，任一匹配即可）
                if period_filter:
                    plist = [p.strip() for p in period.split('/') if p.strip()]
                    if not any(p in period_filter for p in plist):
                        continue
                if teacher_filter and teacher not in teacher_filter:
                    continue

                status = (sd.get('status') or '').strip()
                cf     = sd.get('continue_fall', None)

                if '在读' in status:
                    by_subject[subj]['active'] += 1
                    if pt and pt != '-':
                        by_point[pt] = by_point.get(pt, {'active':0,'refund':0,'renewal':0})
                        by_point[pt]['active'] += 1
                    for p in period.split('/'):
                        p = p.strip()
                        if p and (not period_filter or p in period_filter):
                            by_period[p] = by_period.get(p, 0) + 1
                    if teacher and teacher != '-':
                        by_teacher[teacher] = by_teacher.get(teacher, 0) + 1
                    if _is_cf(cf):
                        by_subject[subj]['renewal'] += 1
                        if pt and pt in by_point:
                            by_point[pt]['renewal'] += 1
                elif status and status != '-' and '在读' not in status:
                    by_subject[subj]['refund'] += 1
                    if pt and pt in by_point:
                        by_point[pt]['refund'] += 1

        new_count = _compute_filtered_new_count(
            conn, rid, point_filter, subj_filter, period_filter, teacher_filter
        )

        active_pv   = sum(v['active'] for v in by_subject.values())
        refund_pv   = sum(v['refund'] for v in by_subject.values())
        renewal_pv  = sum(v['renewal'] for v in by_subject.values())

        trends.append({
            'run_id':     rid,
            'run_time':   rtime,
            'total_uv':   len(snapshots),
            'active_pv':  active_pv,
            'refund_pv':  refund_pv,
            'renewal_pv': renewal_pv,
            'new_count':   new_count,
            'by_subject': by_subject,
            'by_point':   by_point,
            'by_period':  by_period,
            'by_teacher': by_teacher,
        })

    conn.close()
    return json_response({'trends': trends})


@app.route('/api/overview/stats', methods=['GET'])
def api_overview_stats():
    """数据总览统计"""
    latest = get_latest_run()
    if latest is None:
        return json_response({'stats': {}})

    students = latest['students']
    total = len(students)
    active = sum(1 for s in students if '在读' in s.get('status', ''))
    refund = sum(1 for s in students if '退费' in s.get('status', ''))

    # 按教学点统计
    by_point = {}
    for s in students:
        pt = s.get('teaching_point', '未知')
        by_point[pt] = by_point.get(pt, 0) + 1

    # 按学科统计在读人数
    import json as j
    by_subject = {'雪球思维': 0, '悦读创作': 0, '双语素养': 0}
    by_period = {}
    by_teacher = {}

    active_pv = 0
    refund_pv = 0
    renewal_pv = 0
    active_uv = 0
    refund_uv = 0

    # 分学科续报统计
    subject_active_pv = {'雪球思维': 0, '悦读创作': 0, '双语素养': 0}
    subject_renewed_pv = {'雪球思维': 0, '悦读创作': 0, '双语素养': 0}

    # 分学科 × 分期次续报统计
    period_order = get_period_order()
    subject_period_active_pv = {subj: {p: 0 for p in period_order} for subj in subject_active_pv}
    subject_period_renewed_pv = {subj: {p: 0 for p in period_order} for subj in subject_renewed_pv}

    # 全量指标 × 分期次统计
    period_active_pv = {p: 0 for p in period_order}
    period_renewal_pv = {p: 0 for p in period_order}
    period_refund_pv = {p: 0 for p in period_order}
    period_active_uv = {p: set() for p in period_order}
    period_refund_uv = {p: set() for p in period_order}
    period_total_uv = {p: set() for p in period_order}

    for s in students:
        has_active = False
        try:
            subjects = j.loads(s.get('subjects_json', '{}'))
        except:
            subjects = {}
        for subj, subj_data in subjects.items():
            if isinstance(subj_data, dict):
                status = subj_data.get('status', '')
                if '在读' in status:
                    active_pv += 1
                    has_active = True
                    by_subject[subj] = by_subject.get(subj, 0) + 1
                    subject_active_pv[subj] = subject_active_pv.get(subj, 0) + 1
                    period = subj_data.get('period', '')
                    if period and period != '-':
                        for p in period.split('/'):
                            p = p.strip()
                            if p:
                                by_period[p] = by_period.get(p, 0) + 1
                                if p in period_order:
                                    subject_period_active_pv[subj][p] += 1
                                    period_active_pv[p] += 1
                                    period_total_uv[p].add(s.get('uid'))
                    teacher = subj_data.get('teacher', '')
                    if teacher and teacher != '-':
                        by_teacher[teacher] = by_teacher.get(teacher, 0) + 1
                    if is_continue_fall(subj_data.get('continue_fall', '')):
                        renewal_pv += 1
                        subject_renewed_pv[subj] = subject_renewed_pv.get(subj, 0) + 1
                        if period and period != '-':
                            for p in period.split('/'):
                                p = p.strip()
                                if p in period_order:
                                    subject_period_renewed_pv[subj][p] += 1
                                    period_renewal_pv[p] += 1
                # 退款科目：非'-'且非'在读'即为退费（含课前退、课后退等）
                if status and status != '-' and '在读' not in status:
                    refund_pv += 1
                    period = subj_data.get('period', '')
                    if period and period != '-':
                        for p in period.split('/'):
                            p = p.strip()
                            if p in period_order:
                                period_refund_pv[p] += 1
                                period_refund_uv[p].add(s.get('uid'))
                                period_total_uv[p].add(s.get('uid'))
        if has_active:
            active_uv += 1
            period = None
            # 只要学员有任一科目在读，记录其所处期次
            for subj, subj_data in subjects.items():
                if isinstance(subj_data, dict) and '在读' in subj_data.get('status', ''):
                    p_str = subj_data.get('period', '')
                    if p_str and p_str != '-':
                        for p in p_str.split('/'):
                            p = p.strip()
                            if p in period_order:
                                period_active_uv[p].add(s.get('uid'))

    # 退费UV = 三科均不在读的学员
    refund_uv = total - active_uv

    # ── 退费UV期次归属（逐期独立判断）──
    # 规则：对于每个学员的每个期次，如果该学员在该期次有科目但该期次所有科目都退费了，
    #       就计为该期次退费UV。总退费UV仍然是三科全退。
    # 结果：各期退费UV之和 ≥ 总退费UV（一个学员可同时归属多个期次）
    period_refund_uv_final = {p: set() for p in period_order}
    for s in students:
        uid = s.get('uid')
        try:
            subjects = j.loads(s.get('subjects_json', '{}'))
        except Exception:
            subjects = {}
        # 按期次分组：统计每个期次中该学员的科目情况
        period_has_subject = set()       # 该学员涉及的期次
        period_has_active = set()        # 有在读科目的期次
        period_has_refund_only = set()   # 仅有退费科目（无在读）的期次
        for subj, subj_data in subjects.items():
            if isinstance(subj_data, dict):
                status = subj_data.get('status', '')
                period = subj_data.get('period', '')
                if period and period != '-':
                    for p in period.split('/'):
                        p = p.strip()
                        if p in period_order:
                            period_has_subject.add(p)
                            if '在读' in status:
                                period_has_active.add(p)
                            elif status and status != '-':
                                # 退费科目
                                pass  # 只需记录在 period_has_subject
        # 对每个期次：学员在该期次有科目但没有在读 → 计为该期次退费UV
        for p in period_has_subject:
            if p not in period_has_active:
                period_refund_uv_final[p].add(uid)

    total_renewal_rate = round(renewal_pv / active_pv * 100, 1) if active_pv > 0 else 0
    total_pv = active_pv + refund_pv
    total_refund_rate = round(refund_pv / total_pv * 100, 1) if total_pv > 0 else 0

    math_renewal_rate = round(subject_renewed_pv['雪球思维'] / subject_active_pv['雪球思维'] * 100, 1) if subject_active_pv['雪球思维'] > 0 else 0
    chinese_renewal_rate = round(subject_renewed_pv['悦读创作'] / subject_active_pv['悦读创作'] * 100, 1) if subject_active_pv['悦读创作'] > 0 else 0
    english_renewal_rate = round(subject_renewed_pv['双语素养'] / subject_active_pv['双语素养'] * 100, 1) if subject_active_pv['双语素养'] > 0 else 0

    # 分学科 × 分期次续报率
    renewal_rate_by_period = {}
    for subj in subject_active_pv:
        rates = {}
        for p in period_order:
            a = subject_period_active_pv[subj].get(p, 0)
            r = subject_period_renewed_pv[subj].get(p, 0)
            rates[p] = round(r / a * 100, 1) if a > 0 else 0
        renewal_rate_by_period[subj] = rates

    # 按班型统计
    by_class_type = {}
    for s in students:
        try:
            subjects = j.loads(s.get('subjects_json', '{}'))
        except:
            subjects = {}
        for subj, subj_data in subjects.items():
            if isinstance(subj_data, dict):
                ct = subj_data.get('class_type', '')
                if ct and ct != '-':
                    key = f'{subj}-{ct}'
                    by_class_type[key] = by_class_type.get(key, 0) + 1

    # 全量指标 × 分期次汇总
    total_renewal_rate_by_period = {}
    total_refund_rate_by_period = {}
    for p in period_order:
        a = period_active_pv.get(p, 0)
        r = period_renewal_pv.get(p, 0)
        f = period_refund_pv.get(p, 0)
        total_renewal_rate_by_period[p] = round(r / a * 100, 1) if a > 0 else 0
        total_refund_rate_by_period[p] = round(f / (a + f) * 100, 1) if (a + f) > 0 else 0

    period_active_pv_out = {p: period_active_pv.get(p, 0) for p in period_order}
    period_refund_pv_out = {p: period_refund_pv.get(p, 0) for p in period_order}
    period_renewal_pv_out = {p: period_renewal_pv.get(p, 0) for p in period_order}
    period_active_uv_out = {p: len(period_active_uv.get(p, set())) for p in period_order}
    period_refund_uv_out = {p: len(period_refund_uv_final.get(p, set())) for p in period_order}
    period_total_uv_out = {p: len(period_total_uv.get(p, set())) for p in period_order}

    return json_response({
        'total_uv': total,
        'active_uv': active_uv,
        'refund_uv': refund_uv,
        'active_pv': active_pv,
        'refund_pv': refund_pv,
        'renewal_pv': renewal_pv,
        'total_renewal_rate': total_renewal_rate,
        'total_refund_rate': total_refund_rate,
        'math_renewal_rate': math_renewal_rate,
        'chinese_renewal_rate': chinese_renewal_rate,
        'english_renewal_rate': english_renewal_rate,
        'renewal_rate_by_period': renewal_rate_by_period,
        'active_pv_by_period': period_active_pv_out,
        'refund_pv_by_period': period_refund_pv_out,
        'renewal_pv_by_period': period_renewal_pv_out,
        'active_uv_by_period': period_active_uv_out,
        'refund_uv_by_period': period_refund_uv_out,
        'total_uv_by_period': period_total_uv_out,
        'total_renewal_rate_by_period': total_renewal_rate_by_period,
        'total_refund_rate_by_period': total_refund_rate_by_period,
        # 保留旧字段兼容
        'total': total,
        'active': active_uv,
        'refund': refund_uv,
        'by_teaching_point': by_point,
        'by_subject': by_subject,
        'by_period': by_period,
        'by_teacher': by_teacher,
        'by_class_type': by_class_type,
        'run_time': latest['run']['run_time'],
    })


@app.route('/api/overview/export', methods=['POST'])
def api_overview_export():
    """导出筛选后的数据为xlsx"""
    try:
        latest = get_latest_run()
        if latest is None:
            return json_response({'error': '无数据'}, 400)

        run_id = latest['run']['run_id']
        filters = request.get_json() or {}
        students = get_filtered_students(run_id, filters)

        import json as j
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '学员数据'

        headers = ['教学点', 'uid', '姓名', '顾问', '在读状态', '期次',
                   '雪球-班级ID', '雪球-主讲', '雪球-教室', '雪球-时段', '雪球-期次', '雪球-状态',
                   '悦读-班级ID', '悦读-主讲', '悦读-教室', '悦读-时段', '悦读-期次', '悦读-状态',
                   '双语-班级ID', '双语-主讲', '双语-教室', '双语-时段', '双语-期次', '双语-状态']

        header_font = Font(bold=True)
        grey_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = grey_fill

        for i, s in enumerate(students):
            row = 2 + i
            try:
                subjects = j.loads(s.get('subjects_json', '{}'))
            except:
                subjects = {}
            ws.cell(row=row, column=1, value=s.get('teaching_point', ''))
            ws.cell(row=row, column=2, value=s.get('uid', ''))
            ws.cell(row=row, column=3, value=s.get('name', ''))
            ws.cell(row=row, column=4, value=s.get('advisor', ''))
            ws.cell(row=row, column=5, value=s.get('status', ''))
            ws.cell(row=row, column=6, value=s.get('period_combined', ''))

            for si, subj in enumerate(['雪球思维', '悦读创作', '双语素养']):
                sd = subjects.get(subj, {})
                if isinstance(sd, dict):
                    base = 7 + si * 6
                    ws.cell(row=row, column=base, value=sd.get('class_id', ''))
                    ws.cell(row=row, column=base+1, value=sd.get('teacher', ''))
                    ws.cell(row=row, column=base+2, value=sd.get('room', ''))
                    ws.cell(row=row, column=base+3, value=sd.get('time_slot', ''))
                    ws.cell(row=row, column=base+4, value=sd.get('period', ''))
                    ws.cell(row=row, column=base+5, value=sd.get('status', ''))

        out_path = os.path.join(app.config['OUTPUT_FOLDER'], f'筛选导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(out_path)
        wb.close()

        return send_file(out_path, as_attachment=True,
                        download_name=f'UV台帳_筛选导出.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        traceback.print_exc()
        return json_response({'error': str(e)}, 500)


# ── API: 数据总览增强 ──

@app.route('/api/overview/matrix', methods=['GET'])
def api_overview_matrix():
    """期次×教学点矩阵"""
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    matrix_data = get_period_teaching_point_matrix(latest['run']['run_id'])
    return json_response(matrix_data)


@app.route('/api/overview/classes', methods=['GET'])
def api_overview_classes():
    """排班排课看板数据"""
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    pt = request.args.get('teaching_point', '')
    periods_str = request.args.get('periods', '')
    teacher = request.args.get('teacher', '')
    periods = [p.strip() for p in periods_str.split(',') if p.strip()] if periods_str else None
    data = get_class_schedule(latest['run']['run_id'], pt or None, periods, teacher or None)
    return json_response(data)


@app.route('/api/overview/teacher/<teacher>', methods=['GET'])
def api_overview_teacher(teacher):
    """老师维度详情"""
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    period = request.args.get('period', '') or None
    teaching_point = request.args.get('teaching_point', '') or None
    data = get_teacher_detail(latest['run']['run_id'], teacher, period, teaching_point)
    return json_response(data)


@app.route('/api/overview/teacher/<teacher>/export', methods=['GET'])
def api_overview_teacher_export(teacher):
    """导出讲师名下学员名单 xlsx"""
    import json as j
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    period = request.args.get('period', '') or None
    teaching_point = request.args.get('teaching_point', '') or None
    data = get_teacher_detail(latest['run']['run_id'], teacher, period, teaching_point)
    students = data.get('students', [])
    if not students:
        return json_response({'error': '该讲师名下暂无学员'}, 400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{teacher}学员名单'

    # 根据学员学科决定搭班列标题
    first_subj = students[0].get('subject', '')
    co_label_map = {
        '雪球思维': ('搭班语文', '搭班英语', 'chinese_teacher', 'english_teacher'),
        '悦读创作': ('搭班数学', '搭班英语', 'math_teacher', 'english_teacher'),
        '双语素养': ('搭班数学', '搭班语文', 'math_teacher', 'chinese_teacher'),
    }
    co_h1, co_h2, co_k1, co_k2 = co_label_map.get(first_subj, ('搭班语文', '搭班英语', 'chinese_teacher', 'english_teacher'))

    headers = ['校区', '期次', '班级ID', '上课时间', '班型', '学生姓名',
               '主讲', co_h1, co_h2, '顾问老师', '在读情况', '续秋情况']

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    center_align = Alignment(horizontal='center', vertical='center')

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border

    subj_map = {'雪球思维': '数学', '悦读创作': '语文', '双语素养': '英语'}

    for idx, s in enumerate(students):
        row = 2 + idx
        values = [
            s.get('teaching_point', ''),
            s.get('period', ''),
            s.get('class_id', ''),
            s.get('time_slot', ''),
            s.get('class_type', ''),
            s.get('name', ''),
            teacher,
            s.get(co_k1, '') or '-',
            s.get(co_k2, '') or '-',
            s.get('advisor', ''),
            s.get('status', ''),
            s.get('continue_fall', ''),
        ]
        for i, v in enumerate(values, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.alignment = center_align
            c.border = thin_border
            # 在读行绿色，退费行红色
            if '退费' in s.get('status', ''):
                c.font = Font(color='C00000')
            elif '在读' in s.get('status', ''):
                c.font = Font(color='006100')

    # 自动调整列宽
    col_widths = [14, 10, 16, 16, 12, 12, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    out_path = os.path.join(app.config['OUTPUT_FOLDER'],
                            f'讲师_{teacher}_学员名单_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(out_path)
    wb.close()

    return send_file(out_path, as_attachment=True,
                     download_name=f'{teacher}_学员名单.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/overview/teacher-list', methods=['GET'])
def api_overview_teacher_list():
    """老师列表（含简要统计），支持期次/教学点/老师筛选"""
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    period = request.args.get('period', '') or None
    teaching_point = request.args.get('teaching_point', '') or None
    teacher = request.args.get('teacher', '') or None
    data = get_teacher_list(latest['run']['run_id'], period, teaching_point, teacher)
    return json_response({'teachers': data})


@app.route('/api/overview/advisor-list', methods=['GET'])
def api_overview_advisor_list():
    """顾问列表（含转化统计：UV+PV 双维度，支持筛选联动）"""
    import json as j
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    run_id = latest['run']['run_id']

    # 筛选参数：period为空或'all'表示全部数据
    filter_period = request.args.get('period', '')
    filter_pt = request.args.get('teaching_point', '')
    filter_teacher = request.args.get('teacher', '')
    teacher_list = [t.strip() for t in filter_teacher.split(',') if t.strip()] if filter_teacher else []
    all_periods_mode = not filter_period or filter_period == 'all'

    # 判定当前期次（用于非全部数据模式）
    from datetime import datetime as dt_now
    now_dt = dt_now.now()
    auto_period = get_auto_period()

    target_period = filter_period if filter_period and filter_period != 'all' else auto_period

    period_start_str = get_period_start_str(target_period, now_dt.year)

    conn = get_db()

    latest_rows = conn.execute(
        'SELECT uid, name, advisor, teaching_point, subjects_json FROM student_snapshots WHERE run_id = ?',
        (run_id,)
    ).fetchall()

    # 收集筛选选项
    all_teaching_points = set()
    all_teachers = set()

    for r in latest_rows:
        if r['teaching_point']:
            all_teaching_points.add(r['teaching_point'])
        try:
            subs = j.loads(r['subjects_json'] or '{}')
        except Exception:
            continue
        for _subj, _sd in subs.items():
            if isinstance(_sd, dict):
                _t = _sd.get('teacher', '')
                if _t:
                    all_teachers.add(_t)

    conn.close()

    advisor_stats = {}
    advisor_students = {}  # {advisor: [{uid,name,...}]}

    for r in latest_rows:
        uid = r['uid']

        # 教学点筛选
        if filter_pt and r['teaching_point'] != filter_pt:
            continue

        advisor = r['advisor'] or '未知'
        if not advisor or advisor == '-':
            advisor = '未知'
        try:
            subjects = j.loads(r['subjects_json'] or '{}')
        except Exception:
            continue

        has_active_in_period = False
        active_subjects = []
        renewed_subjects = []
        student_subjects_info = []
        pre_renewed_subjs = set()  # all_periods_mode: 已续科目
        new_renewed_subjs = set()  # all_periods_mode: 新增续费科目

        for _subj, _sd in subjects.items():
            if not isinstance(_sd, dict):
                continue
            # 老师筛选：只看匹配老师的科目
            if teacher_list and _sd.get('teacher', '') not in teacher_list:
                continue
            _status = _sd.get('status', '-')
            _period_val = _sd.get('period', '') or ''
            _period_list = [p.strip() for p in _period_val.split('/') if p.strip()]
            _is_active = '在读' in _status
            _is_renewed = is_continue_fall(_sd.get('continue_fall', ''))
            student_subjects_info.append({
                'subject': _subj,
                'status': _status,
                'period': _period_val,
                'class_id': _sd.get('class_id', ''),
                'teacher': _sd.get('teacher', ''),
                'time_slot': _sd.get('time_slot', '-'),
                'is_active': _is_active,
                'is_renewed': _is_renewed,
                'continue_fall': _sd.get('continue_fall', ''),
            })
            if all_periods_mode:
                # 全部数据模式：只要该科目在读即计入
                if _is_active:
                    has_active_in_period = True
                    active_subjects.append(_subj)
                if _is_renewed:
                    renewed_subjects.append(_subj)
                    # 用秋季续费支付时间判断 pre/new
                    subj_period = _period_list[0] if _period_list else ''
                    renewal_class = _classify_renewal_by_pay_time(
                        _sd.get('fall_pay_time'), subj_period, now_dt.year)
                    if renewal_class == 'pre':
                        pre_renewed_subjs.add(_subj)
                    else:
                        new_renewed_subjs.add(_subj)
            else:
                if target_period in _period_list:
                    if _is_active:
                        has_active_in_period = True
                        active_subjects.append(_subj)
                    if _is_renewed:
                        renewed_subjects.append(_subj)
                        renewal_class = _classify_renewal_by_pay_time(
                            _sd.get('fall_pay_time'), target_period, now_dt.year)
                        if renewal_class == 'pre':
                            pre_renewed_subjs.add(_subj)
                        else:
                            new_renewed_subjs.add(_subj)

        if not has_active_in_period:
            continue

        if advisor not in advisor_stats:
            advisor_stats[advisor] = {
                'uv': {'active': set(), 'pre': set(), 'new': set()},
                'pv': {'active': 0, 'pre': 0, 'new': 0}
            }
            advisor_students[advisor] = []

        advisor_stats[advisor]['uv']['active'].add(uid)
        is_renewed_now = len(renewed_subjects) > 0
        uv_renew_type = None
        if is_renewed_now:
            # 有任意"当期转化"科目 → new；全部为"开课前续费" → pre
            if new_renewed_subjs:
                advisor_stats[advisor]['uv']['new'].add(uid)
                uv_renew_type = 'new'
            else:
                advisor_stats[advisor]['uv']['pre'].add(uid)
                uv_renew_type = 'pre'

        for _subj in active_subjects:
            advisor_stats[advisor]['pv']['active'] += 1
            if _subj in renewed_subjects:
                if _subj in new_renewed_subjs:
                    advisor_stats[advisor]['pv']['new'] += 1
                else:
                    advisor_stats[advisor]['pv']['pre'] += 1

        advisor_students[advisor].append({
            'uid': uid,
            'name': r['name'],
            'advisor': advisor,
            'subjects': student_subjects_info,
            'active_subjects': active_subjects,
            'renewed_subjects': renewed_subjects,
            'renew_type': uv_renew_type,
            'current_period': '' if all_periods_mode else target_period,
        })

    advisor_list = []
    for advisor, sets in advisor_stats.items():
        uv_active = len(sets['uv']['active'])
        uv_pre = len(sets['uv']['pre'])
        uv_new = len(sets['uv']['new'])
        uv_should = uv_active - uv_pre
        uv_rate = round(uv_new / uv_should * 100, 1) if uv_should > 0 else 0
        pv_active = sets['pv']['active']
        pv_pre = sets['pv']['pre']
        pv_new = sets['pv']['new']
        pv_should = pv_active - pv_pre
        pv_rate = round(pv_new / pv_should * 100, 1) if pv_should > 0 else 0
        uv_total = uv_pre + uv_new   # 已转化 = 开课前续报 + 当期转化
        pv_total = pv_pre + pv_new   # PV已转化
        advisor_list.append({
            'name': advisor,
            'active': uv_active,
            'pre_renewed': uv_pre,
            'should_renew': uv_should,
            'new_renewed': uv_new,
            'total_renewed': uv_total,
            'renewal_rate': uv_rate,
            'pv_active': pv_active,
            'pv_pre_renewed': pv_pre,
            'pv_should_renew': pv_should,
            'pv_new_renewed': pv_new,
            'pv_total_renewed': pv_total,
            'pv_renewal_rate': pv_rate,
            'student_count': len(advisor_students.get(advisor, [])),
        })
    advisor_list.sort(key=lambda x: x['renewal_rate'], reverse=True)

    return json_response({
        'advisors': advisor_list,
        'current_period': '' if all_periods_mode else target_period,
        'auto_period': auto_period,
        'all_periods_mode': all_periods_mode,
        'teaching_points': sorted(all_teaching_points),
        'teachers': sorted(all_teachers),
        'students_by_advisor': {k: v for k, v in advisor_students.items()},
    })


@app.route('/api/overview/advisor/<path:advisor>', methods=['GET'])
def api_overview_advisor_detail(advisor):
    """顾问维度详情：名下学员列表（支持筛选联动）"""
    import json as j
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    run_id = latest['run']['run_id']

    # 筛选参数：period为空或'all'表示全部数据
    filter_period = request.args.get('period', '')
    filter_pt = request.args.get('teaching_point', '')
    filter_teacher = request.args.get('teacher', '')
    teacher_list = [t.strip() for t in filter_teacher.split(',') if t.strip()] if filter_teacher else []
    all_periods_mode = not filter_period or filter_period == 'all'

    # 判定当前期次（用于非全部数据模式）
    from datetime import datetime as dt_now
    now_dt = dt_now.now()
    auto_period = get_auto_period()

    current_period = filter_period if filter_period and filter_period != 'all' else auto_period

    period_start_str = get_period_start_str(current_period, now_dt.year)

    conn = get_db()

    if advisor == '未知':
        rows = conn.execute(
            'SELECT uid, name, teaching_point, advisor, subjects_json, manual_json '
            'FROM student_snapshots WHERE run_id = ? '
            'AND (advisor IS NULL OR advisor = "" OR advisor = "-")',
            (run_id,)
        ).fetchall()
    else:
        rows = conn.execute(
            'SELECT uid, name, teaching_point, advisor, subjects_json, manual_json '
            'FROM student_snapshots WHERE run_id = ? AND advisor = ?',
            (run_id, advisor)
        ).fetchall()
    conn.close()

    students = []
    subj_map = {'雪球思维': '数', '悦读创作': '语', '双语素养': '英'}
    for r in rows:
        # 教学点筛选
        if filter_pt and r['teaching_point'] != filter_pt:
            continue
        try:
            subjects = j.loads(r['subjects_json'] or '{}')
        except Exception:
            continue
        has_active_in_period = False
        active_subjects = []
        renewed_subjects = []
        subject_details = []
        pre_renewed_subjs = set()  # all_periods_mode
        new_renewed_subjs = set()  # all_periods_mode
        uid_ref = r['uid']
        for _subj, _sd in subjects.items():
            if not isinstance(_sd, dict):
                continue
            # 老师筛选
            if teacher_list and _sd.get('teacher', '') not in teacher_list:
                continue
            _status = _sd.get('status', '-')
            _period_val = _sd.get('period', '') or ''
            _period_list = [p.strip() for p in _period_val.split('/') if p.strip()]
            _is_active = '在读' in _status
            _is_renewed = is_continue_fall(_sd.get('continue_fall', ''))
            if all_periods_mode:
                in_current = _is_active
                if _is_active:
                    has_active_in_period = True
                    active_subjects.append(_subj)
                if _is_renewed:
                    renewed_subjects.append(_subj)
                    # 用秋季续费支付时间判断 pre/new
                    subj_period = _period_list[0] if _period_list else ''
                    renewal_class = _classify_renewal_by_pay_time(
                        _sd.get('fall_pay_time'), subj_period, now_dt.year)
                    if renewal_class == 'pre':
                        pre_renewed_subjs.add(_subj)
                    else:
                        new_renewed_subjs.add(_subj)
            else:
                in_current = current_period in _period_list
                if in_current and _is_active:
                    has_active_in_period = True
                    active_subjects.append(_subj)
                if in_current and _is_renewed:
                    renewed_subjects.append(_subj)
                    renewal_class = _classify_renewal_by_pay_time(
                        _sd.get('fall_pay_time'), current_period, now_dt.year)
                    if renewal_class == 'pre':
                        pre_renewed_subjs.add(_subj)
                    else:
                        new_renewed_subjs.add(_subj)
            subject_details.append({
                'subject': _subj,
                'subject_short': subj_map.get(_subj, _subj[:1]),
                'status': _status,
                'period': _period_val,
                'class_id': _sd.get('class_id', ''),
                'teacher': _sd.get('teacher', ''),
                'time_slot': _sd.get('time_slot', '-'),
                'is_active': _is_active,
                'is_renewed': _is_renewed,
                'in_current_period': in_current,
                'continue_fall': _sd.get('continue_fall', ''),
            })

        if not has_active_in_period:
            continue

        uid = uid_ref
        is_renewed_now = len(renewed_subjects) > 0
        renew_type = None
        if is_renewed_now:
            if new_renewed_subjs:
                renew_type = 'new'
            else:
                renew_type = 'pre'

        # PV per subject
        pv_active = len([s for s in subject_details if s['in_current_period'] and s['is_active']])
        pv_renewed_new = 0
        pv_renewed_pre = 0
        for sd in subject_details:
            if sd['in_current_period'] and sd['is_active'] and sd['is_renewed']:
                if sd['subject'] in new_renewed_subjs:
                    pv_renewed_new += 1
                else:
                    pv_renewed_pre += 1

        students.append({
            'uid': uid,
            'name': r['name'],
            'teaching_point': r['teaching_point'],
            'subject_details': subject_details,
            'active_subjects': [subj_map.get(s, s[:1]) for s in active_subjects],
            'renewed_subjects': [subj_map.get(s, s[:1]) for s in renewed_subjects],
            'renew_type': renew_type,
            'pv_active': pv_active,
            'pv_renewed_new': pv_renewed_new,
            'pv_renewed_pre': pv_renewed_pre,
        })

    return json_response({
        'advisor': advisor,
        'current_period': '' if all_periods_mode else current_period,
        'all_periods_mode': all_periods_mode,
        'total_students': len(students),
        'students': students,
    })


@app.route('/api/overview/student/<uid>', methods=['GET'])
def api_overview_student(uid):
    """学员详情卡片"""
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    data = get_student_detail(latest['run']['run_id'], uid)
    if data is None:
        return json_response({'error': '学员不存在'}, 404)
    return json_response(data)


@app.route('/api/share-card', methods=['GET'])
def api_share_card():
    """分享卡片数据：按期次/教学点/学科/老师筛选，返回KPI+分教学点+分学科+老师排行"""
    import json as j

    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无数据'}, 400)

    run_id = latest['run']['run_id']

    # 筛选参数
    period = request.args.get('period', '')       # 零期/一期/二期/三期
    point = request.args.get('teaching_point', '')  # 教学点
    subject = request.args.get('subject', '')       # 学科
    teacher = request.args.get('teacher', '')       # 老师
    advisor_filter = request.args.get('advisor', '')  # 顾问

    ALL_SUBJS = ['雪球思维', '悦读创作', '双语素养']
    subj_filter = subject if subject else None
    if subj_filter and subj_filter not in ALL_SUBJS:
        subj_filter = None

    conn = get_db()
    rows = conn.execute(
        'SELECT uid, name, teaching_point, advisor, subjects_json FROM student_snapshots WHERE run_id = ?',
        (run_id,)
    ).fetchall()
    conn.close()

    # 遍历筛选
    total_active = 0
    total_renewed = 0
    total_refund = 0
    total_uv = set()
    total_renewed_uv = set()
    total_classes = set()

    by_point = {}     # {pt: {active, renewed, refund, uv:set, renewed_uv:set, classes:set}}
    by_subject = {s: {'active': 0, 'renewed': 0, 'refund': 0} for s in ALL_SUBJS}
    by_teacher = {}   # {teacher: {active, renewed, refund}}
    by_class = {}     # {class_id: {active, renewed, uv:set, renewed_uv:set, subject, period}}

    for r in rows:
        uid = r['uid']
        pt = r['teaching_point'] or ''

        # 顾问筛选
        r_advisor = r['advisor'] or '未知'
        if not r_advisor or r_advisor == '-':
            r_advisor = '未知'
        if advisor_filter and r_advisor != advisor_filter:
            continue

        try:
            subjects = j.loads(r['subjects_json'] or '{}')
        except Exception:
            continue

        for subj_name, sd in subjects.items():
            if not isinstance(sd, dict):
                continue
            if subj_filter and subj_name != subj_filter:
                continue

            status = sd.get('status', '-')
            period_val = sd.get('period', '') or ''
            teacher_val = sd.get('teacher', '') or ''
            class_id = sd.get('class_id', '') or ''

            # 期次筛选
            if period:
                plist = [p.strip() for p in period_val.split('/') if p.strip()]
                if period not in plist:
                    continue

            # 教学点筛选
            if point and pt != point:
                continue

            # 老师筛选
            if teacher and teacher_val != teacher:
                continue

            is_active = '在读' in status
            is_renewed = is_continue_fall(sd.get('continue_fall', ''))
            is_refund = status and status != '-' and '在读' not in status

            if is_active:
                total_active += 1
                total_uv.add(uid)
                if class_id and class_id != '-':
                    total_classes.add(class_id)

                if pt not in by_point:
                    by_point[pt] = {'active': 0, 'renewed': 0, 'refund': 0, 'uv': set(), 'renewed_uv': set(), 'classes': set()}
                by_point[pt]['active'] += 1
                by_point[pt]['uv'].add(uid)
                if class_id and class_id != '-':
                    by_point[pt]['classes'].add(class_id)

                if subj_name in by_subject:
                    by_subject[subj_name]['active'] += 1

                if teacher_val and teacher_val != '-':
                    if teacher_val not in by_teacher:
                        by_teacher[teacher_val] = {'active': 0, 'renewed': 0, 'refund': 0}
                    by_teacher[teacher_val]['active'] += 1

                # 班级维度（仅当筛选了老师时才跟踪）
                if teacher and class_id and class_id != '-':
                    if class_id not in by_class:
                        by_class[class_id] = {'active': 0, 'renewed': 0, 'uv': set(), 'renewed_uv': set(), 'subject': subj_name, 'period': period_val}
                    by_class[class_id]['active'] += 1
                    by_class[class_id]['uv'].add(uid)

                if is_renewed:
                    total_renewed += 1
                    total_renewed_uv.add(uid)
                    if pt in by_point:
                        by_point[pt]['renewed'] += 1
                        by_point[pt]['renewed_uv'].add(uid)
                    if subj_name in by_subject:
                        by_subject[subj_name]['renewed'] += 1
                    if teacher_val and teacher_val != '-' and teacher_val in by_teacher:
                        by_teacher[teacher_val]['renewed'] += 1
                    if class_id in by_class:
                        by_class[class_id]['renewed'] += 1
                        by_class[class_id]['renewed_uv'].add(uid)

            if is_refund:
                total_refund += 1
                if pt in by_point:
                    by_point[pt]['refund'] += 1
                if subj_name in by_subject:
                    by_subject[subj_name]['refund'] += 1
                if teacher_val and teacher_val != '-' and teacher_val in by_teacher:
                    by_teacher[teacher_val]['refund'] += 1

    # 提取全部老师/教学点/顾问（不受筛选影响，用于前端chip选择器）
    all_teachers = set()
    all_points = set()
    all_advisors = set()
    for r in rows:
        pt_all = r['teaching_point'] or ''
        if pt_all and pt_all != '-':
            all_points.add(pt_all)
        adv_all = r['advisor'] or '未知'
        if adv_all and adv_all != '-':
            all_advisors.add(adv_all)
        try:
            subjects_all = j.loads(r['subjects_json'] or '{}')
        except Exception:
            subjects_all = {}
        for subj_name, sd in subjects_all.items():
            if not isinstance(sd, dict):
                continue
            tv = sd.get('teacher', '') or ''
            if tv and tv != '-':
                all_teachers.add(tv)

    # 组装返回数据
    renewal_rate = round(total_renewed / total_active * 100, 1) if total_active > 0 else 0
    renewal_uv = len(total_renewed_uv)
    uv_count = len(total_uv)
    renewal_uv_rate = round(renewal_uv / uv_count * 100, 1) if uv_count > 0 else 0

    # 分教学点表格
    point_list = []
    for pt, v in sorted(by_point.items()):
        if not pt:
            continue
        pt_active = v['active']
        pt_renewed = v['renewed']
        pt_uv = len(v['uv'])
        pt_renewed_uv = len(v['renewed_uv'])
        pt_rate = round(pt_renewed / pt_active * 100, 1) if pt_active > 0 else 0
        pt_uv_rate = round(pt_renewed_uv / pt_uv * 100, 1) if pt_uv > 0 else 0
        point_list.append({
            'name': pt,
            'active': pt_active,
            'uv': pt_uv,
            'renewed': pt_renewed,
            'renewed_uv': pt_renewed_uv,
            'refund': v['refund'],
            'class_count': len(v['classes']),
            'renewal_rate': pt_rate,
            'renewal_uv_rate': pt_uv_rate,
        })

    # 分学科
    subject_list = []
    subj_labels = {'雪球思维': '数学', '悦读创作': '语文', '双语素养': '英语'}
    subj_colors = {'雪球思维': '#378ADD', '悦读创作': '#639922', '双语素养': '#BA7517'}
    for s in ALL_SUBJS:
        v = by_subject[s]
        rate = round(v['renewed'] / v['active'] * 100, 1) if v['active'] > 0 else 0
        subject_list.append({
            'name': subj_labels.get(s, s),
            'key': s,
            'active': v['active'],
            'renewed': v['renewed'],
            'refund': v['refund'],
            'renewal_rate': rate,
            'color': subj_colors.get(s, '#888'),
        })

    # 老师排行
    teacher_list = []
    for t, v in by_teacher.items():
        if not t or t == '-':
            continue
        rate = round(v['renewed'] / v['active'] * 100, 1) if v['active'] > 0 else 0
        teacher_list.append({
            'name': t,
            'active': v['active'],
            'renewed': v['renewed'],
            'renewal_rate': rate,
        })
    teacher_list.sort(key=lambda x: x['renewal_rate'], reverse=True)

    # 班级续报详情（仅当筛选了老师时返回）
    class_list = []
    if teacher:
        subj_labels_full = {'雪球思维': '数学', '悦读创作': '语文', '双语素养': '英语'}
        for cid, v in sorted(by_class.items(), key=lambda x: x[1]['active'], reverse=True):
            c_active = v['active']
            c_renewed = v['renewed']
            c_uv = len(v['uv'])
            c_renewed_uv = len(v['renewed_uv'])
            c_rate = round(c_renewed / c_active * 100, 1) if c_active > 0 else 0
            c_uv_rate = round(c_renewed_uv / c_uv * 100, 1) if c_uv > 0 else 0
            class_list.append({
                'class_id': cid,
                'subject': subj_labels_full.get(v['subject'], v['subject']),
                'period': v['period'] or '-',
                'active': c_active,
                'uv': c_uv,
                'renewed': c_renewed,
                'renewed_uv': c_renewed_uv,
                'renewal_rate': c_rate,
                'renewal_uv_rate': c_uv_rate,
            })

    # ── 顾问维度（联动筛选：period/point/subject/teacher/advisor）──
    # 使用秋季续费订单的支付时间判断 pre/new，无需快照基线
    from datetime import datetime as dt_now
    now_dt = dt_now.now()
    current_period = get_auto_period()

    all_periods_mode = not period  # 无期次筛选 → 全部期次
    target_period_adv = period if period else current_period

    conn2 = get_db()

    # 加载最新 run（含 advisor + teaching_point 字段）
    latest_rows = conn2.execute(
        'SELECT uid, advisor, teaching_point, subjects_json FROM student_snapshots WHERE run_id = ?', (run_id,)
    ).fetchall()
    conn2.close()

    # 按顾问聚合（UV + PV 双维度），应用 point/subject/teacher/advisor 筛选
    advisor_stats = {}
    for r in latest_rows:
        uid = r['uid']
        if point and (r['teaching_point'] or '') != point:
            continue

        advisor = r['advisor'] or '未知'
        if not advisor or advisor == '-':
            advisor = '未知'
        if advisor_filter and advisor != advisor_filter:
            continue

        try:
            subjects = j.loads(r['subjects_json'] or '{}')
        except Exception:
            continue

        has_active = False
        active_subjs = []
        renewed_subjs = []
        pre_renewed_subjs = set()
        new_renewed_subjs = set()

        for _subj, _sd in subjects.items():
            if not isinstance(_sd, dict):
                continue
            if subj_filter and _subj != subj_filter:
                continue
            if teacher and _sd.get('teacher', '') != teacher:
                continue
            _status = _sd.get('status', '-')
            _period_val = _sd.get('period', '') or ''
            _period_list = [p.strip() for p in _period_val.split('/') if p.strip()]
            _is_active = '在读' in _status
            _is_renewed = is_continue_fall(_sd.get('continue_fall', ''))

            if all_periods_mode:
                # 全部期次模式：只要在读就计入
                if _is_active:
                    has_active = True
                    active_subjs.append(_subj)
                if _is_renewed:
                    renewed_subjs.append(_subj)
                    # 用秋季续费支付时间判断 pre/new
                    subj_period = _period_list[0] if _period_list else ''
                    renewal_class = _classify_renewal_by_pay_time(
                        _sd.get('fall_pay_time'), subj_period, now_dt.year)
                    if renewal_class == 'pre':
                        pre_renewed_subjs.add(_subj)
                    else:
                        new_renewed_subjs.add(_subj)
            else:
                # 单期模式：只看目标期次
                if target_period_adv in _period_list:
                    if _is_active:
                        has_active = True
                        active_subjs.append(_subj)
                    if _is_renewed:
                        renewed_subjs.append(_subj)
                        renewal_class = _classify_renewal_by_pay_time(
                            _sd.get('fall_pay_time'), target_period_adv, now_dt.year)
                        if renewal_class == 'pre':
                            pre_renewed_subjs.add(_subj)
                        else:
                            new_renewed_subjs.add(_subj)

        if not has_active:
            continue

        if advisor not in advisor_stats:
            advisor_stats[advisor] = {
                'uv': {'active': set(), 'pre': set(), 'new': set()},
                'pv': {'active': 0, 'pre': 0, 'new': 0}
            }

        # UV: uid 级别
        advisor_stats[advisor]['uv']['active'].add(uid)
        if new_renewed_subjs:
            advisor_stats[advisor]['uv']['new'].add(uid)
        elif renewed_subjs:
            advisor_stats[advisor]['uv']['pre'].add(uid)

        # PV: (uid, subject) 级别
        for _subj in active_subjs:
            advisor_stats[advisor]['pv']['active'] += 1
            if _subj in renewed_subjs:
                if _subj in new_renewed_subjs:
                    advisor_stats[advisor]['pv']['new'] += 1
                else:
                    advisor_stats[advisor]['pv']['pre'] += 1

    advisor_list = []
    for advisor, sets in advisor_stats.items():
        uv_active = len(sets['uv']['active'])
        uv_pre = len(sets['uv']['pre'])
        uv_new = len(sets['uv']['new'])
        uv_should = uv_active - uv_pre
        uv_rate = round(uv_new / uv_should * 100, 1) if uv_should > 0 else 0
        pv_active = sets['pv']['active']
        pv_pre = sets['pv']['pre']
        pv_new = sets['pv']['new']
        pv_should = pv_active - pv_pre
        pv_rate = round(pv_new / pv_should * 100, 1) if pv_should > 0 else 0
        uv_total = uv_pre + uv_new
        pv_total = pv_pre + pv_new
        advisor_list.append({
            'name': advisor,
            'active': uv_active,
            'pre_renewed': uv_pre,
            'should_renew': uv_should,
            'new_renewed': uv_new,
            'total_renewed': uv_total,
            'renewal_rate': uv_rate,
            'pv_active': pv_active,
            'pv_pre_renewed': pv_pre,
            'pv_should_renew': pv_should,
            'pv_new_renewed': pv_new,
            'pv_total_renewed': pv_total,
            'pv_renewal_rate': pv_rate,
        })
    advisor_list.sort(key=lambda x: x['renewal_rate'], reverse=True)

    return json_response({
        'kpi': {
            'active_pv': total_active,
            'uv': uv_count,
            'renewal_rate': renewal_rate,
            'renewal_pv': total_renewed,
            'renewal_uv': renewal_uv,
            'renewal_uv_rate': renewal_uv_rate,
            'refund_pv': total_refund,
            'class_count': len(total_classes),
        },
        'by_point': point_list,
        'by_subject': subject_list,
        'by_teacher': teacher_list,
        'by_class': class_list,
        'by_advisor': advisor_list,
        'advisor_current_period': '全部' if all_periods_mode else target_period_adv,
        'filters': {
            'period': period or '全部',
            'teaching_point': point or '全部',
            'subject': subj_labels.get(subj_filter, subj_filter) if subj_filter else '全部',
            'teacher': teacher or '全部',
            'advisor': advisor_filter or '全部',
        },
        'available_teachers': sorted(all_teachers),
        'available_points': sorted(all_points),
        'available_advisors': sorted(all_advisors),
        'run_time': latest['run']['run_time'],
    })


# ═══ 分享页面生成 ═══

SHARES_FOLDER = os.path.join(DATA_DIR, 'shares')
os.makedirs(SHARES_FOLDER, exist_ok=True)


@app.route('/share-view/<share_id>')
def view_share(share_id):
    """本地预览生成的分享页面"""
    share_dir = os.path.join(SHARES_FOLDER, share_id)
    index_path = os.path.join(share_dir, 'index.html')
    if not os.path.exists(index_path):
        return 'Share not found', 404
    return send_file(index_path)


@app.route('/api/share/generate', methods=['POST'])
def api_share_generate():
    """
    生成分享页面：获取最新数据，渲染自包含HTML，保存到shares/目录。
    支持全部6个筛选参数：start_date, end_date, subjects, teaching_points, periods, teachers。
    返回 share_id，前端可通过 /share-view/<share_id> 本地预览。
    """
    import json as j
    from datetime import datetime as dt_now_class, timedelta

    data = request.get_json() or {}

    # ── 0. 解析筛选参数（与 api_trends_data 保持一致）──
    start_date = (data.get('start_date') or '').strip()
    end_date   = (data.get('end_date') or '').strip()

    def _parse_filter(val):
        """将逗号分隔字符串或列表解析为去空列表"""
        if not val:
            return []
        if isinstance(val, list):
            return [x.strip() for x in val if str(x).strip()]
        return [x.strip() for x in str(val).split(',') if x.strip()]

    subjects_filter = _parse_filter(data.get('subjects', ''))
    points_filter   = _parse_filter(data.get('teaching_points', ''))
    periods_filter  = _parse_filter(data.get('periods', ''))
    teachers_filter = _parse_filter(data.get('teachers', ''))

    # 向后兼容：旧参数 period / teaching_point（单值）
    old_period = (data.get('period') or '').strip()
    old_point  = (data.get('teaching_point') or '').strip()
    if old_period and not periods_filter:
        periods_filter = [old_period]
    if old_point and not points_filter:
        points_filter = [old_point]

    # 保存用户选择的期次，仅用于分享页默认Tab和筛选文案；
    # 分享页数据始终包含全部期次，以便在外链中切换查看。
    selected_periods = periods_filter
    periods_filter = []

    now = dt_now_class.now()
    today_str = now.strftime('%Y-%m-%d')

    # ── 1. 获取最新快照，计算 KPI / 主讲排行 / 顾问排行 ──
    latest = get_latest_run()
    if latest is None:
        return json_response({'error': '无校准数据'}, 400)
    run_id = latest['run']['run_id']
    run_time = latest['run']['run_time'] or today_str

    conn = get_db()
    rows = conn.execute(
        'SELECT uid, name, teaching_point, advisor, subjects_json FROM student_snapshots WHERE run_id = ?',
        (run_id,)
    ).fetchall()

    ALL_SUBJS = ['雪球思维', '悦读创作', '双语素养']

    total_active = 0
    total_renewed = 0
    total_refund = 0
    total_uv = set()
    total_renewed_uv = set()

    by_teacher_pv = {}   # {teacher: {active, renewed, refund, pre, new, subjects}}
    advisor_stats = {}   # {advisor: {uv:{active,set, pre:set, new:set}, pv:{...}}}

    PERIODS = get_period_order()
    by_teacher_period = {p: {} for p in PERIODS}   # {period: {teacher: {active, renewed, pre, new, subjects}}}
    by_advisor_period  = {p: {} for p in PERIODS}   # {period: {advisor: {active, pre, new, uv_active, uv_pre, uv_new}}}
    uv_by_period = {p: set() for p in PERIODS}      # 每期次在读学员UV
    renewed_uv_by_period = {p: set() for p in PERIODS}  # 每期次续秋学员UV

    for r in rows:
        uid = r['uid']
        pt = (r['teaching_point'] or '').strip()
        advisor = r['advisor'] or '未知'
        if not advisor or advisor == '-':
            advisor = '未知'

        # 教学点筛选
        if points_filter and pt not in points_filter:
            continue

        try:
            subjects = j.loads(r['subjects_json'] or '{}')
        except Exception:
            continue

        uid_has_active = False
        uid_active_pv_count = 0  # PV维度：本uid的活跃学科数
        uid_pre_set = set()
        uid_new_set = set()

        for subj_name, sd in subjects.items():
            if not isinstance(sd, dict):
                continue

            # 学科筛选
            if subjects_filter and subj_name not in subjects_filter:
                continue

            status     = sd.get('status', '-')
            period_val = sd.get('period', '') or ''
            teacher_val = sd.get('teacher', '') or ''

            # 期次筛选（跨期用/拼接，任一匹配即可）
            if periods_filter:
                plist = [p.strip() for p in period_val.split('/') if p.strip()]
                if not any(p in periods_filter for p in plist):
                    continue

            # 主讲筛选
            if teachers_filter and teacher_val not in teachers_filter:
                continue

            is_active  = '在读' in status
            is_renewed = is_continue_fall(sd.get('continue_fall', ''))
            is_refund  = status and status != '-' and '在读' not in status

            if is_active:
                total_active += 1
                total_uv.add(uid)
                uid_has_active = True
                uid_active_pv_count += 1

                # 按期次统计在读UV
                for pp in [p.strip() for p in period_val.split('/') if p.strip()]:
                    if pp in uv_by_period:
                        uv_by_period[pp].add(uid)

                # 主讲 PV
                if teacher_val and teacher_val != '-':
                    if teacher_val not in by_teacher_pv:
                        by_teacher_pv[teacher_val] = {'active': 0, 'renewed': 0, 'refund': 0, 'pre': 0, 'new': 0, 'subjects': set()}
                    by_teacher_pv[teacher_val]['active'] += 1
                    by_teacher_pv[teacher_val]['subjects'].add(subj_name)

                    # 按期次跟踪主讲（仅统计筛选内的期次）
                    for pp in [p.strip() for p in period_val.split('/') if p.strip()]:
                        if pp in by_teacher_period and (not periods_filter or pp in periods_filter):
                            if teacher_val not in by_teacher_period[pp]:
                                by_teacher_period[pp][teacher_val] = {'active': 0, 'renewed': 0, 'pre': 0, 'new': 0, 'subjects': set()}
                            by_teacher_period[pp][teacher_val]['active'] += 1
                            by_teacher_period[pp][teacher_val]['subjects'].add(subj_name)

                if is_renewed:
                    total_renewed += 1
                    total_renewed_uv.add(uid)
                    if teacher_val and teacher_val != '-' and teacher_val in by_teacher_pv:
                        by_teacher_pv[teacher_val]['renewed'] += 1

                    # 按期次统计续秋UV
                    for pp in [p.strip() for p in period_val.split('/') if p.strip()]:
                        if pp in renewed_uv_by_period:
                            renewed_uv_by_period[pp].add(uid)

                    # 按期次跟踪主讲续报
                    for pp in [p.strip() for p in period_val.split('/') if p.strip()]:
                        if pp in by_teacher_period and (not periods_filter or pp in periods_filter):
                            if teacher_val and teacher_val != '-':
                                if teacher_val in by_teacher_period[pp]:
                                    by_teacher_period[pp][teacher_val]['renewed'] += 1

                    # Pre/new 分类
                    if periods_filter:
                        subj_periods = [p.strip() for p in period_val.split('/') if p.strip()]
                        matching = [p for p in periods_filter if p in subj_periods]
                        classify_period = matching[0] if matching else periods_filter[0]
                    else:
                        classify_period = _detect_subj_period(period_val)
                    renewal_class = _classify_renewal_by_pay_time(
                        sd.get('fall_pay_time'), classify_period, now.year)
                    if renewal_class == 'pre':
                        uid_pre_set.add(subj_name)
                    else:
                        uid_new_set.add(subj_name)
                    # 主讲 PV 级别 pre/new 统计
                    if teacher_val and teacher_val != '-' and teacher_val in by_teacher_pv:
                        if renewal_class == 'pre':
                            by_teacher_pv[teacher_val]['pre'] += 1
                        else:
                            by_teacher_pv[teacher_val]['new'] += 1
                    # 按期次主讲 pre/new 统计
                    for pp in [p.strip() for p in period_val.split('/') if p.strip()]:
                        if pp in by_teacher_period and (not periods_filter or pp in periods_filter):
                            if teacher_val and teacher_val != '-' and teacher_val in by_teacher_period[pp]:
                                if renewal_class == 'pre':
                                    by_teacher_period[pp][teacher_val]['pre'] += 1
                                else:
                                    by_teacher_period[pp][teacher_val]['new'] += 1

            if is_refund:
                total_refund += 1
                if teacher_val and teacher_val != '-' and teacher_val in by_teacher_pv:
                    by_teacher_pv[teacher_val]['refund'] += 1

        # 聚合顾问 UV + PV（每位学员计入一次）
        if uid_has_active:
            if advisor not in advisor_stats:
                advisor_stats[advisor] = {
                    'uv': {'active': set(), 'pre': set(), 'new': set()},
                    'pv': {'active': 0, 'pre': 0, 'new': 0},
                }
            advisor_stats[advisor]['uv']['active'].add(uid)
            if uid_new_set:
                advisor_stats[advisor]['uv']['new'].add(uid)
            elif uid_pre_set:
                advisor_stats[advisor]['uv']['pre'].add(uid)
            # PV维度：(uid, subject) 级别计数
            advisor_stats[advisor]['pv']['active'] += uid_active_pv_count
            advisor_stats[advisor]['pv']['new'] += len(uid_new_set)
            advisor_stats[advisor]['pv']['pre'] += len(uid_pre_set)

            # 按期次跟踪顾问（去重：每位学员在每期次最多计一次）
            advisor_period_seen = set()
            for subj_name, sd in subjects.items():
                if not isinstance(sd, dict):
                    continue
                # 学科筛选（与主循环一致）
                if subjects_filter and subj_name not in subjects_filter:
                    continue
                pv = sd.get('period', '') or ''
                status = sd.get('status', '-')
                if '在读' not in status:
                    continue
                # 主讲筛选
                teacher_val = sd.get('teacher', '') or ''
                if teachers_filter and teacher_val not in teachers_filter:
                    continue
                for pp in [p.strip() for p in pv.split('/') if p.strip()]:
                    if pp not in by_advisor_period:
                        continue
                    # 期次筛选：仅统计筛选内的期次
                    if periods_filter and pp not in periods_filter:
                        continue
                    key = f'{pp}|{advisor}'
                    if key in advisor_period_seen:
                        continue
                    advisor_period_seen.add(key)
                    if advisor not in by_advisor_period[pp]:
                        by_advisor_period[pp][advisor] = {
                            'active': 0, 'pre': 0, 'new': 0,
                            'uv_active': set(), 'uv_pre': set(), 'uv_new': set(),
                        }
                    by_advisor_period[pp][advisor]['active'] += 1
                    by_advisor_period[pp][advisor]['uv_active'].add(uid)
                    if uid_new_set:
                        by_advisor_period[pp][advisor]['new'] += 1
                        by_advisor_period[pp][advisor]['uv_new'].add(uid)
                    elif uid_pre_set:
                        by_advisor_period[pp][advisor]['pre'] += 1
                        by_advisor_period[pp][advisor]['uv_pre'].add(uid)

    # ── 主讲排行（PV维度，含续报率+转化率）──
    teacher_list = []
    for tname, v in sorted(by_teacher_pv.items()):
        active = v['active']
        renewed = v['renewed']
        pre = v.get('pre', 0)
        new = v.get('new', 0)
        if active == 0:
            continue
        rate = round(renewed / active * 100, 1)
        should = active - pre
        conv_rate = round(new / should * 100, 1) if should > 0 else 0
        teacher_list.append({
            'name': tname,
            'active': active,
            'renewed': renewed,
            'renewal_rate': rate,
            'pre_renewed': pre,
            'should_renew': should,
            'new_renewed': new,
            'total_renewed': pre + new,
            'conv_rate': conv_rate,
            'subjects': sorted(v.get('subjects', set())),
        })
    teacher_list.sort(key=lambda x: x['renewal_rate'], reverse=True)

    # ── 顾问排行（UV+PV双维度）──
    advisor_list_raw = []
    for aname, sets in advisor_stats.items():
        uv_active = len(sets['uv']['active'])
        uv_pre = len(sets['uv']['pre'])
        uv_new = len(sets['uv']['new'])
        uv_should = uv_active - uv_pre
        uv_total = uv_pre + uv_new
        pv_active = sets['pv']['active']
        pv_pre = sets['pv']['pre']
        pv_new = sets['pv']['new']
        pv_should = pv_active - pv_pre
        pv_total = pv_pre + pv_new
        conv_rate = round(uv_new / uv_should * 100, 1) if uv_should > 0 else 0
        if uv_active == 0:
            continue
        advisor_list_raw.append({
            'name': aname,
            'active': uv_active,
            'pre_renewed': uv_pre,
            'should_renew': uv_should,
            'new_renewed': uv_new,
            'total_renewed': uv_total,
            'conv_rate': conv_rate,
            'pv_active': pv_active,
            'pv_pre_renewed': pv_pre,
            'pv_should_renew': pv_should,
            'pv_new_renewed': pv_new,
            'pv_total_renewed': pv_total,
        })
    advisor_list_raw.sort(key=lambda x: x['conv_rate'], reverse=True)

    # ── 按期次主讲排行 ──
    teachers_by_period = {}
    for pp in PERIODS:
        tlist = []
        for tname, v in by_teacher_period[pp].items():
            active = v['active']
            renewed = v.get('renewed', 0)
            pre = v.get('pre', 0)
            new = v.get('new', 0)
            if active == 0:
                continue
            rate = round(renewed / active * 100, 1)
            should = active - pre
            conv_rate = round(new / should * 100, 1) if should > 0 else 0
            tlist.append({
                'name': tname, 'active': active, 'renewed': renewed,
                'renewal_rate': rate,
                'pre_renewed': pre, 'should_renew': should,
                'new_renewed': new, 'total_renewed': pre + new,
                'conv_rate': conv_rate,
                'subjects': sorted(v.get('subjects', set())),
            })
        tlist.sort(key=lambda x: x['renewal_rate'], reverse=True)
        teachers_by_period[pp] = tlist

    # ── 按期次顾问排行 ──
    advisors_by_period = {}
    for pp in PERIODS:
        alist = []
        for aname, v in by_advisor_period[pp].items():
            uv_active = len(v.get('uv_active', set()))
            uv_pre = len(v.get('uv_pre', set()))
            uv_new = len(v.get('uv_new', set()))
            uv_should = uv_active - uv_pre
            pv_active = v['active']
            pv_pre = v.get('pre', 0)
            pv_new = v.get('new', 0)
            pv_should = pv_active - pv_pre
            if uv_active == 0:
                continue
            conv_rate = round(uv_new / uv_should * 100, 1) if uv_should > 0 else 0
            alist.append({
                'name': aname,
                'active': uv_active,
                'pre_renewed': uv_pre,
                'new_renewed': uv_new,
                'should_renew': uv_should,
                'total_renewed': uv_pre + uv_new,
                'conv_rate': conv_rate,
                'pv_active': pv_active,
                'pv_pre_renewed': pv_pre,
                'pv_should_renew': pv_should,
                'pv_new_renewed': pv_new,
                'pv_total_renewed': pv_pre + pv_new,
            })
        alist.sort(key=lambda x: x['conv_rate'], reverse=True)
        advisors_by_period[pp] = alist

    # ── KPI ──
    renewal_rate = round(total_renewed / total_active * 100, 1) if total_active > 0 else 0
    uv_count = len(total_uv)
    renewal_uv_count = len(total_renewed_uv)
    renewal_uv_rate = round(renewal_uv_count / uv_count * 100, 1) if uv_count > 0 else 0

    kpis = {
        'active_pv': total_active,
        'refund_pv': total_refund,
        'renewal_pv': total_renewed,
        'uv_count': uv_count,
        'renewal_uv_count': renewal_uv_count,
        'renewal_rate': renewal_rate,
        'renewal_uv_rate': renewal_uv_rate,
    }

    # ── 2. 趋势数据（按天、按期次聚合，含 new / new_renew）──
    # 构建带日期筛选的 SQL（与 api_trends_data 逻辑一致）
    trend_sql = 'SELECT * FROM calibration_runs WHERE 1=1'
    trend_params = []
    if start_date:
        trend_sql += ' AND date(run_time) >= ?'
        trend_params.append(start_date)
    if end_date:
        trend_sql += ' AND date(run_time) <= ?'
        trend_params.append(end_date)
    if not start_date and not end_date:
        # 默认近7天
        default_start = (now - timedelta(days=7)).strftime('%Y-%m-%d')
        trend_sql += ' AND date(run_time) >= ?'
        trend_params.append(default_start)
    trend_sql += ' ORDER BY run_time ASC'

    trend_runs = conn.execute(trend_sql, trend_params).fetchall()

    # 按天取最后一次校准
    day_map = {}
    for row in trend_runs:
        run = dict(row)
        day = run['run_time'][:10]
        if day not in day_map or run['run_time'] > day_map[day]['run_time']:
            day_map[day] = run

    trend_all = []
    trend_by_period = {p: [] for p in PERIODS}

    # 用于计算 new_renew 的前一日 renew 值
    prev_renew_all = 0
    prev_renew_period = {p: 0 for p in PERIODS}

    for day in sorted(day_map.keys()):
        run = day_map[day]
        rid = run['run_id']
        snapshots = conn.execute(
            'SELECT * FROM student_snapshots WHERE run_id = ?', (rid,)
        ).fetchall()

        day_active = 0
        day_refund = 0
        day_renew  = 0
        pd = {p: {'active': 0, 'refund': 0, 'renew': 0} for p in PERIODS}

        for s_row in snapshots:
            s = dict(s_row)
            s_pt = (s.get('teaching_point') or '').strip()
            # 教学点筛选
            if points_filter and s_pt not in points_filter:
                continue

            try:
                subs = j.loads(s.get('subjects_json') or '{}')
            except Exception:
                subs = {}

            for subj, sd in subs.items():
                if not isinstance(sd, dict):
                    continue
                # 学科筛选
                if subjects_filter and subj not in subjects_filter:
                    continue

                status     = (sd.get('status') or '').strip()
                period_val = (sd.get('period') or '').strip()
                cf         = sd.get('continue_fall', None)
                teacher_val = (sd.get('teacher') or '').strip()

                # 期次筛选（跨期任一匹配即可）
                if periods_filter:
                    plist = [p.strip() for p in period_val.split('/') if p.strip()]
                    if not any(p in periods_filter for p in plist):
                        continue

                # 主讲筛选
                if teachers_filter and teacher_val not in teachers_filter:
                    continue

                if '在读' in status:
                    day_active += 1
                    if _is_cf(cf):
                        day_renew += 1

                    # 按具体期次分配（仅统计筛选内的期次）
                    for p in period_val.split('/'):
                        p = p.strip()
                        if p not in pd:
                            continue
                        if periods_filter and p not in periods_filter:
                            continue
                        pd[p]['active'] += 1
                        if _is_cf(cf):
                            pd[p]['renew'] += 1

                elif status and status != '-' and '在读' not in status:
                    day_refund += 1
                    for p in period_val.split('/'):
                        p = p.strip()
                        if p not in pd:
                            continue
                        if periods_filter and p not in periods_filter:
                            continue
                        pd[p]['refund'] += 1

        # 新增学员数（按筛选条件联动：教学点/学科/期次/主讲）
        new_count = _compute_filtered_new_count(
            conn, rid, points_filter, subjects_filter, periods_filter, teachers_filter
        )

        # 当日新增续费 = max(0, 今日续费 - 昨日续费）
        new_renew_all = max(0, day_renew - prev_renew_all)
        new_renew_pd = {}
        for p in PERIODS:
            new_renew_pd[p] = max(0, pd[p]['renew'] - prev_renew_period[p])

        # 更新前一日值
        prev_renew_all = day_renew
        for p in PERIODS:
            prev_renew_period[p] = pd[p]['renew']

        date_label = day[-5:]  # MM-DD
        trend_all.append({
            'date': date_label,
            'active': day_active,
            'refund': day_refund,
            'renew': day_renew,
            'new': new_count,
            'new_renew': new_renew_all,
        })
        for p in PERIODS:
            trend_by_period[p].append({
                'date': date_label,
                'active': pd[p]['active'],
                'refund': pd[p]['refund'],
                'renew': pd[p]['renew'],
                'new': new_count,           # 新增学员不按期次拆分，与 all 一致
                'new_renew': new_renew_pd[p],
            })

    conn.close()

    trend_data = {'all': trend_all}
    for p in PERIODS:
        trend_data[p] = trend_by_period[p]

    # 有数据的期次列表（用于前端控制Tab显示）
    periods_available = ['all']
    for p in PERIODS:
        has_data = any(
            d['active'] > 0 or d['refund'] > 0 or d['renew'] > 0 or d['new_renew'] > 0
            for d in trend_by_period[p]
        )
        if has_data:
            periods_available.append(p)

    # ── 2.5 按期次拆分的 KPI（最后一天的快照即当前累计）──
    kpis_by_period = {}
    for p in PERIODS:
        if trend_by_period[p]:
            last = trend_by_period[p][-1]
            active = last['active']
            renew = last['renew']
            refund = last['refund']
            uv_active = len(uv_by_period.get(p, set()))
            uv_renew = len(renewed_uv_by_period.get(p, set()))
            kpis_by_period[p] = {
                'active_pv': active,
                'refund_pv': refund,
                'renewal_pv': renew,
                'uv_count': uv_active,
                'renewal_uv_count': uv_renew,
                'renewal_rate': round(renew / active * 100, 1) if active > 0 else 0,
                'renewal_uv_rate': round(uv_renew / uv_active * 100, 1) if uv_active > 0 else 0,
            }
        else:
            kpis_by_period[p] = {
                'active_pv': 0,
                'refund_pv': 0,
                'renewal_pv': 0,
                'uv_count': 0,
                'renewal_uv_count': 0,
                'renewal_rate': 0,
                'renewal_uv_rate': 0,
            }

    # ── 3. 筛选文本（包含全部已激活的筛选条件）──
    filter_parts = []
    if start_date:
        filter_parts.append(f'开始: {start_date}')
    if end_date:
        filter_parts.append(f'结束: {end_date}')
    if subjects_filter:
        filter_parts.append(f'学科: {"/".join(subjects_filter)}')
    if points_filter:
        filter_parts.append(f'校区: {"/".join(points_filter)}')
    if selected_periods:
        filter_parts.append(f'期次: {"/".join(selected_periods)}')
    if teachers_filter:
        filter_parts.append(f'主讲: {"/".join(teachers_filter)}')
    filter_text = ' · '.join(filter_parts) if filter_parts else '全部数据'

    # default_period：单期次筛选时联动到该Tab，否则显示"全部"
    default_period = selected_periods[0] if len(selected_periods) == 1 else '全部'

    # ── 4. 渲染模板 ──
    share_id = uuid.uuid4().hex[:12]
    share_dir = os.path.join(SHARES_FOLDER, share_id)
    os.makedirs(share_dir, exist_ok=True)

    html = render_template(
        'share_page.html',
        update_time=today_str,
        trend_data_json=j.dumps(trend_data, ensure_ascii=False),
        kpis_json=j.dumps(kpis, ensure_ascii=False),
        kpis_by_period_json=j.dumps(kpis_by_period, ensure_ascii=False),
        teachers_json=j.dumps(teacher_list, ensure_ascii=False),
        advisors_json=j.dumps(advisor_list_raw, ensure_ascii=False),
        teachers_by_period_json=j.dumps(teachers_by_period, ensure_ascii=False),
        advisors_by_period_json=j.dumps(advisors_by_period, ensure_ascii=False),
        periods_available_json=j.dumps(periods_available, ensure_ascii=False),
        filter_text=filter_text,
        default_period=default_period,
        period_order=get_period_order(),
    )

    index_path = os.path.join(share_dir, 'index.html')
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(html)

    # 读取该分享目录已部署的外部URL（如果有），避免返回旧缓存
    deploy_url = ''
    share_deploy_meta = os.path.join(share_dir, '.deploy_url')
    if os.path.exists(share_deploy_meta):
        with open(share_deploy_meta, 'r', encoding='utf-8') as f:
            deploy_url = f.read().strip()

    # 兼容旧全局缓存：如果该目录没有单独部署过，再尝试读全局缓存
    if not deploy_url:
        global_deploy_meta = os.path.join(SHARES_FOLDER, '.deploy_url')
        if os.path.exists(global_deploy_meta):
            with open(global_deploy_meta, 'r', encoding='utf-8') as f:
                deploy_url = f.read().strip()

    return json_response({
        'share_id': share_id,
        'local_url': f'/share-view/{share_id}',
        'share_dir': share_dir,
        'deploy_url': deploy_url,
        'kpis': kpis,
        'teacher_count': len(teacher_list),
        'advisor_count': len(advisor_list_raw),
        'trend_days': len(trend_all),
        'periods_available': periods_available,
    })


def _detect_subj_period(period_val):
    """从跨期字符串中提取主要期次"""
    if not period_val:
        return ''
    parts = [p.strip() for p in period_val.split('/') if p.strip()]
    return parts[0] if parts else ''


def _is_cf(v):
    """判断续秋标记"""
    if not v:
        return False
    v = str(v).strip()
    return v in ('是', '已续', 'T', 'AC', 'AL', 'true', '1', '√')


@app.route('/api/share/deploy', methods=['POST'])
def api_share_deploy():
    """请求部署分享页面到外网。写入请求文件，由管理员(WorkBuddy)处理部署。
    返回 request_id，前端可轮询 /api/share/deploy-status/<request_id> 获取结果。
    """
    data = request.get_json() or {}
    share_dir = data.get('share_dir', '')
    share_id = data.get('share_id', '')

    if not share_dir or not os.path.isdir(share_dir):
        return json_response({'error': '分享目录不存在，请先生成分享页面'}, 400)

    deploy_requests_dir = os.path.join(SHARES_FOLDER, '.deploy_requests')
    os.makedirs(deploy_requests_dir, exist_ok=True)

    request_id = uuid.uuid4().hex[:10]
    request_file = os.path.join(deploy_requests_dir, f'{request_id}.json')

    req_data = {
        'request_id': request_id,
        'share_dir': share_dir,
        'share_id': share_id or os.path.basename(share_dir),
        'created_at': datetime.now().isoformat(),
        'status': 'pending',
    }
    with open(request_file, 'w', encoding='utf-8') as f:
        f.write(json.dumps(req_data, ensure_ascii=False))

    return json_response({
        'request_id': request_id,
        'status': 'pending',
        'message': '已加入部署队列，请稍候...',
    })


@app.route('/api/share/deploy-status/<request_id>', methods=['GET'])
def api_share_deploy_status(request_id):
    """查询部署状态"""
    deploy_requests_dir = os.path.join(SHARES_FOLDER, '.deploy_requests')
    request_file = os.path.join(deploy_requests_dir, f'{request_id}.json')

    if not os.path.exists(request_file):
        return json_response({'error': '请求不存在'}, 404)

    with open(request_file, 'r', encoding='utf-8') as f:
        req_data = json.loads(f.read())

    status = req_data.get('status', 'pending')
    url = req_data.get('url', None)

    return json_response({
        'request_id': request_id,
        'status': status,
        'url': url,
    })


if __name__ == '__main__':
    PORT = 5100
    if getattr(sys, 'frozen', False):
        # 打包模式：自动打开浏览器
        def open_browser():
            webbrowser.open(f'http://localhost:{PORT}')
        threading.Timer(1.5, open_browser).start()
        print(f'UV台帐管理系统(1.0)已启动，浏览器将自动打开 http://localhost:{PORT}')
        print(f'数据目录: {DATA_DIR}')
        print('按 Ctrl+C 或关闭此窗口退出')
    print(f'UV台帐管理系统(1.0) 启动端口: {PORT}')
    app.run(debug=False, port=PORT, host='0.0.0.0')
