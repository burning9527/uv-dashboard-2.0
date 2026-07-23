"""
UV台帳 校准引擎
用全量招生订单校准学科台帐，识别变动并生成输出
"""
import copy
import io
from datetime import datetime
from collections import defaultdict, OrderedDict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 配置
# ============================================================

PERIOD_DATES = {
    '零期': datetime(2026, 6, 29),
    '一期': datetime(2026, 7, 13),
    '二期': datetime(2026, 7, 27),
    '三期': datetime(2026, 8, 10),
}

PERIOD_ORDER = {'零期': 0, '一期': 1, '二期': 2, '三期': 3}

SUBJECT_COLUMNS = {
    '雪球思维': 'J-T',
    '悦读创作': 'U-AC',
    '双语素养': 'AD-AL',
}

SUBJECT_COL_RANGES = {
    '雪球思维': (10, 20),   # J=10, T=20
    '悦读创作': (21, 29),   # U=21, AC=29
    '双语素养': (30, 38),   # AD=30, AL=38
}

# 单科字段映射: (招生明细列名, UV台帐列号, 类型)
ORDER_FIELD_MAP = {
    '雪球思维': {
        '班级id':       ('在读班级id', 10),
        '主讲':         ('主讲', 14),
        '教室':         ('教室', 15),
        '上课时段':     ('上课时段', 16),
        '班型':         ('班型', 17),
        '期次':         ('期次', 18),
        '是否在读':     ('在班状态', 19),
        '是否续秋':     (None, 20),
    },
    '悦读创作': {
        '班级id':       ('在读班级id', 21),
        '主讲':         ('主讲', 23),
        '教室':         ('教室', 24),
        '上课时段':     ('上课时段', 25),
        '班型':         ('班型', 26),
        '期次':         ('期次', 27),
        '是否在读':     ('在班状态', 28),
        '是否续秋':     (None, 29),
    },
    '双语素养': {
        '班级id':       ('在读班级id', 30),
        '主讲':         ('主讲', 32),
        '教室':         ('教室', 33),
        '上课时段':     ('上课时段', 34),
        '班型':         ('班型', 35),
        '期次':         ('期次', 36),
        '是否在读':     ('在班状态', 37),
        '是否续秋':     (None, 38),
    },
}

# 旧台帐学科列映射: subject → {字段: 列号}
OLD_SUBJECT_COL_MAP = {
    '雪球思维': {
        'class_id': 10, 'teacher': 14, 'room': 15,
        'time_slot': 16, 'class_type': 17, 'period': 18, 'status': 19,
    },
    '悦读创作': {
        'class_id': 21, 'teacher': 23, 'room': 24,
        'time_slot': 25, 'class_type': 26, 'period': 27, 'status': 28,
    },
    '双语素养': {
        'class_id': 30, 'teacher': 32, 'room': 33,
        'time_slot': 34, 'class_type': 35, 'period': 36, 'status': 37,
    },
}

CHANGE_LABELS = {
    'class_id': '班级变更',
    'teacher': '主讲变更',
    'room': '教室变更',
    'time_slot': '时段变更',
    'class_type': '班型变更',
    'period': '期次变更',
}


def normalize_teaching_point(raw):
    """标准化教学点名称：去'教学点'后缀；天成保留'大厦'，其他去'大厦'后缀"""
    if not raw:
        return raw
    name = str(raw).strip()
    name = name.removesuffix('教学点').strip()
    if name != '天成大厦':
        name = name.removesuffix('大厦').strip()
    return name


# 已知城市列表（用于从分校字段提取城市名）；按需扩展
_KNOWN_CITIES = (
    '北京', '上海', '广州', '深圳', '天津', '杭州', '南京', '成都', '武汉',
    '西安', '重庆', '苏州', '长沙', '青岛', '宁波', '郑州', '济南', '合肥',
    '福州', '厦门', '南昌', '昆明', '沈阳', '大连', '哈尔滨', '长春',
    '石家庄', '太原', '南宁', '贵阳', '海口', '兰州', '佛山', '东莞',
    '无锡', '常州', '珠海', '中山', '惠州',
)


def normalize_branch(raw):
    """从分校字段提取城市名，面向全国数据。
    例：'K9广州分校' → '广州'；'深圳校区' → '深圳'；'上海城市教学中心' → '上海'。
    无法匹配已知城市时，去除 K9/分校/校区/中心 等词缀兜底。
    """
    if not raw:
        return ''
    s = str(raw).strip()
    for city in _KNOWN_CITIES:
        if city in s:
            return city
    # 兜底：去掉常见前缀/后缀，保留剩余中文字符
    s2 = s
    for token in ('K9', '分校', '校区', '城市', '教学中心', '中心', '校区教学点'):
        s2 = s2.replace(token, '')
    s2 = s2.strip(' -_/')
    return s2


def normalize_season(raw):
    """标准化学季：寒假/暑假/春季/秋季（兼容 '暑' / '寒假班' 等写法）"""
    if not raw:
        return ''
    s = str(raw).strip()
    for kw, norm in (('寒', '寒假'), ('暑', '暑假'), ('春', '春季'), ('秋', '秋季')):
        if kw in s:
            return norm
    return s


def _get_uv_sheet(wb):
    """兼容两种繁简写法，找到 UV台帐/UV台帳 工作表"""
    for name in wb.sheetnames:
        if 'UV台帐' in name or 'UV台帳' in name:
            return wb[name]
    return None


def parse_order_date(date_val):
    """解析订单中的日期字段"""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, str):
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d']:
            try:
                return datetime.strptime(date_val.strip(), fmt)
            except ValueError:
                continue
    return None


def classify_refund(refund_time, subject, period_field_name):
    """
    退费分类：课前退 / 课后退
    按对应学科期次的开课首日判断
    """
    if not refund_time:
        return '课前退'
    rt = parse_order_date(refund_time)
    if rt is None:
        return '课前退'

    period = str(period_field_name).strip() if period_field_name else ''
    for p_name, p_date in PERIOD_DATES.items():
        if p_name in period:
            if rt < p_date:
                return '课前退'
            else:
                return '课后退'
    return '课前退'


def _normalize_status(status_str):
    """标准化状态字符串，用于对比"""
    if not status_str:
        return '-'
    s = str(status_str).strip()
    if s == '-' or s == '':
        return '-'
    if '在读' in s:
        return '在读'
    if '课前退' in s:
        return '课前退'
    if '课后退' in s:
        return '课后退'
    if '退费' in s:
        return '退费'
    return s


def _is_continue_fall(val):
    """判断任意形式的续秋标记是否表示已续报：支持布尔、是/已续/T/AC/AL/1 等"""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if not s or s == '-':
        return False
    return s in {'是', '已续', 't', 'ac', 'al', 'true', 'yes', '1', 'y'}


def _fmt_continue_fall(val):
    """把续秋标记统一为 '是'/'否'/'-' 字符串"""
    if _is_continue_fall(val):
        return '是'
    if val is None or (isinstance(val, str) and (val.strip() == '' or val.strip() == '-')):
        return '-'
    return '否'


# ============================================================
# 数据加载
# ============================================================

def load_enrollment_orders(filepath):
    """
    加载招生明细全量订单
    返回: list of dict
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[col] = str(val).strip()

    orders = []
    for row in range(2, ws.max_row + 1):
        def g(col):
            return ws.cell(row=row, column=col).value

        biz = str(g(1) or '')
        season = str(g(22) or '')
        grade = str(g(23) or '')
        subject = str(g(24) or '')
        # 全国数据：放开业务线/学季/年级硬过滤，仅保留产品学科口径
        if subject not in ('雪球思维', '悦读创作', '双语素养'):
            continue
        # 分校为必需维度，无法识别分校的订单跳过（避免污染全国数据）
        branch = normalize_branch(g(3))
        if not branch:
            continue

        order = {
            'uid': str(g(4) or '').strip(),
            'name': str(g(5) or '').strip(),
            'subject': subject,
            'class_id': str(g(6) or '').strip(),
            'class_name': str(g(7) or '').strip(),
            'period': str(g(12) or '').strip(),
            'teacher': str(g(9) or '').strip(),
            'time_slot': str(g(11) or '').strip(),
            'room': str(g(17) or '').strip(),
            'class_type': str(g(25) or '').strip(),
            'teaching_point_raw': str(g(18) or '').strip(),
            'teaching_point': normalize_teaching_point(g(18)),
            'status': str(g(31) or '').strip(),
            'refund_time': g(37),
            'pay_time': parse_order_date(g(36)),
            'advisor': str(g(41) or '').strip(),
            'start_date': str(g(13) or '').strip(),
            'is_refund': str(g(32) or '').strip() != '',
            'branch': branch,
            'grade': grade.strip(),
            'season': normalize_season(season),
        }
        orders.append(order)

    wb.close()
    return orders


def _fix_known_anomalies(orders):
    """
    修正已知的招生明细数据异常，返回 (fixed_orders, anomaly_records)
    当前已知异常:
    - 李奕萱(4815971458): 三科联报退了数学和语文，但数学退费订单在系统中误标为英语(双语素养)，
      导致出现两个英语订单(一退一在读)、缺少数学订单。修正：将英语退费订单改标记为数学(雪球思维)
    """
    anomalies = []
    fixed = list(orders)

    # ── 李奕萱：英语退费误标为数学退费 ──
    lyt_uid = '4815971458'
    lyt_orders = [o for o in fixed if o['uid'] == lyt_uid]
    if lyt_orders:
        # 按学科分组，找重复学科
        from collections import Counter
        subj_counts = Counter(o['subject'] for o in lyt_orders)
        dup_subjs = [s for s, c in subj_counts.items() if c > 1]

        if dup_subjs:
            # 找到每个重复学科中的退费订单
            for dup_subj in dup_subjs:
                dup_orders = [o for o in lyt_orders if o['subject'] == dup_subj]
                refund_orders = [o for o in dup_orders if o['is_refund']]
                active_orders = [o for o in dup_orders if not o['is_refund']]

                if refund_orders and active_orders:
                    # 检查是否有缺失学科
                    all_subs = {'雪球思维', '悦读创作', '双语素养'}
                    present_subs = set(o['subject'] for o in lyt_orders)
                    missing_subs = all_subs - present_subs

                    if missing_subs:
                        missing_subj = list(missing_subs)[0]
                        # 李奕萱数学退费班级的实际信息
                        math_refund_info = {
                            'subject': '雪球思维',
                            'class_id': '3132880',
                            'period': '零期',
                            'teacher': '陈中帅',
                            'room': '教室2',
                            'time_slot': '15:50-18:00',
                            'class_type': '启能班',
                        }
                        for ro in refund_orders:
                            ro.update(math_refund_info)
                            anomalies.append({
                                'uid': lyt_uid,
                                'name': lyt_orders[0].get('name', '李奕萱'),
                                'desc': f'订单学科异常修正：{dup_subj}退费订单误标为数学退费（班{math_refund_info["class_id"]}，'
                                        f'{math_refund_info["period"]}/{math_refund_info["teacher"]}/{math_refund_info["time_slot"]}）'
                                        f'，原{dup_subj}保留在读订单',
                                'fix': f'{dup_subj}→{missing_subj}（班{math_refund_info["class_id"]} 零期/启能/陈中帅）',
                            })

    return fixed, anomalies


def load_old_ledger_subject_data(filepath):
    """
    从原UV台帐提取每个学员各科的历史数据
    返回: dict[uid] = {subject: {class_id, teacher, room, time_slot, class_type, period, status}}
    """
    if not filepath:
        return {}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _get_uv_sheet(wb)
    if ws is None:
        wb.close()
        return {}

    data = {}
    data_start = 3

    for row in range(data_start, ws.max_row + 1):
        uid_val = ws.cell(row=row, column=2).value
        if not uid_val:
            continue
        uid = str(uid_val).strip()

        data[uid] = {}
        for subject, col_map in OLD_SUBJECT_COL_MAP.items():
            subject_data = {}
            for field, col in col_map.items():
                val = ws.cell(row=row, column=col).value
                subject_data[field] = str(val).strip() if val else '-'
            # 标准化状态
            subject_data['status_normalized'] = _normalize_status(subject_data.get('status', '-'))
            data[uid][subject] = subject_data

    wb.close()
    return data


def load_old_ledger_manual_fields(filepath):
    """
    从原UV台帐提取手动字段映射 uid → {E, F, G, K, L, M, V, AE, AM, AN-AQ, AR, AS, 是否续秋}
    """
    if not filepath:
        return {}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = _get_uv_sheet(wb)
    if ws is None:
        wb.close()
        return {}

    data_start = 3
    manual = {}

    for row in range(data_start, ws.max_row + 1):
        uid_val = ws.cell(row=row, column=2).value
        if not uid_val:
            continue
        uid = str(uid_val).strip()

        manual[uid] = {
            'E': ws.cell(row=row, column=5).value,
            'F': ws.cell(row=row, column=6).value,
            'G': ws.cell(row=row, column=7).value,
            'K': ws.cell(row=row, column=11).value,
            'L': ws.cell(row=row, column=12).value,
            'M': ws.cell(row=row, column=13).value,
            'T': ws.cell(row=row, column=20).value,
            'V': ws.cell(row=row, column=22).value,
            'AC': ws.cell(row=row, column=29).value,
            'AE': ws.cell(row=row, column=31).value,
            'AL': ws.cell(row=row, column=38).value,
            'AM': ws.cell(row=row, column=39).value,
            'AN': ws.cell(row=row, column=40).value,
            'AO': ws.cell(row=row, column=41).value,
            'AP': ws.cell(row=row, column=42).value,
            'AQ': ws.cell(row=row, column=43).value,
            'AR': ws.cell(row=row, column=44).value,
            'AS': ws.cell(row=row, column=45).value,
        }

    wb.close()
    return manual


# ============================================================
# 校准逻辑
# ============================================================

def calibrate(enrollment_file, old_ledger_file, fall_orders_file=None):
    """
    执行台帐校准 — 对比新旧数据，识别各类变动
    返回 dict:
      - students: list of calibrated student dicts
      - changes: list of per-student change records (学员维度)
      - change_details: flat list of detailed change records
      - stats: summary statistics
      - refund_details: only NEW refund details
    """
    orders = load_enrollment_orders(enrollment_file)

    # 修正已知数据异常
    orders, anomaly_records = _fix_known_anomalies(orders)

    old_manual = load_old_ledger_manual_fields(old_ledger_file) if old_ledger_file else {}
    old_subject = load_old_ledger_subject_data(old_ledger_file) if old_ledger_file else {}

    fall_data = {}
    if fall_orders_file:
        fall_data = load_fall_orders(fall_orders_file)

    # 按uid聚合订单
    uid_orders = defaultdict(list)
    for o in orders:
        uid_orders[o['uid']].append(o)

    # ── 构建校准后学员列表 + 逐学员变动检测 ──
    students = []
    student_changes = []       # 学员维度: 每个学员的变动摘要
    change_details = []         # 逐条明细
    all_refund_details = []     # 所有退费明细
    new_refund_details = []     # 仅新增退费

    all_uids = set(uid_orders.keys())
    old_uids = set(old_manual.keys()) | set(old_subject.keys())

    for uid, user_orders in uid_orders.items():
        first = user_orders[0]
        is_new = uid not in old_uids

        student = {
            'uid': uid,
            'name': first['name'],
            'teaching_point': first['teaching_point'],
            'advisor': first.get('advisor', ''),
            'is_new': is_new,
            'branch': first.get('branch', ''),
            'grade': first.get('grade', ''),
            'season': first.get('season', ''),
        }

        # 手动字段继承
        manual = old_manual.get(uid, {})
        student['channel'] = manual.get('E', '')
        student['purpose'] = manual.get('F', '')
        student['notes'] = manual.get('G', '')

        # 各科订单
        subject_orders = defaultdict(list)
        for o in user_orders:
            subject_orders[o['subject']].append(o)

        active_subjects = []
        period_set = set()
        this_student_changes = []      # 该学员的变动

        for subject in ['雪球思维', '悦读创作', '双语素养']:
            s_orders = sorted(
                subject_orders.get(subject, []),
                key=lambda x: (x['pay_time'] or datetime(2000, 1, 1)),
                reverse=True
            )
            latest = s_orders[0] if s_orders else None

            if latest is None:
                # 未报名
                student[f'{subject}_class_id'] = '-'
                student[f'{subject}_teacher'] = '-'
                student[f'{subject}_room'] = '-'
                student[f'{subject}_time_slot'] = '-'
                student[f'{subject}_class_type'] = '-'
                student[f'{subject}_period_single'] = '-'
                student[f'{subject}_status'] = '-'
                student[f'{subject}_continue_fall'] = '-'
                continue

            is_refund = latest['is_refund'] or latest['status'] not in ('', '在读')
            refund_type = None
            if is_refund and latest['refund_time']:
                refund_type = classify_refund(latest['refund_time'], subject, latest['period'])
            elif is_refund:
                refund_type = '课前退'

            # 获取原台帐中该科旧状态
            old_subj = old_subject.get(uid, {}).get(subject, {})
            old_status_norm = old_subj.get('status_normalized', '-')

            new_status = refund_type if refund_type else '在读'

            # 写入学生数据
            student[f'{subject}_class_id'] = latest['class_id']
            student[f'{subject}_teacher'] = latest['teacher']
            student[f'{subject}_room'] = latest['room']
            student[f'{subject}_time_slot'] = latest['time_slot']
            student[f'{subject}_class_type'] = latest['class_type']
            student[f'{subject}_period_single'] = latest['period']
            student[f'{subject}_status'] = new_status

            if is_refund:
                rd = {
                    'uid': uid, 'name': first['name'], 'subject': subject,
                    'type': refund_type, 'time': latest['refund_time'],
                    'period': latest['period']
                }
                all_refund_details.append(rd)

                # 判断是否新增退费
                is_new_refund = (old_status_norm == '-' or old_status_norm == '在读' or uid not in old_uids)
                if is_new_refund:
                    new_refund_details.append(rd)
                    this_student_changes.append({
                        'type': f'新增退费-{subject}',
                        'detail': f'{subject} {refund_type}（原: {old_status_norm} → 现: {refund_type}）',
                        'field': 'status',
                        'old': old_status_norm,
                        'new': refund_type,
                        'subject': subject,
                        'period': latest['period'],
                        'teacher': latest['teacher'],
                    })
            else:
                active_subjects.append(subject)
                period_set.add(latest['period'])

                # 对比原台帐数据检查变动
                if uid in old_uids:
                    for field, label in CHANGE_LABELS.items():
                        old_val = str(old_subj.get(field, '-')).strip()
                        new_val_map = {
                            'class_id': latest['class_id'],
                            'teacher': latest['teacher'],
                            'room': latest['room'],
                            'time_slot': latest['time_slot'],
                            'class_type': latest['class_type'],
                            'period': latest['period'],
                        }
                        new_val = str(new_val_map.get(field, '')).strip()

                        if old_val and old_val != '-' and new_val and new_val != '-' and old_val != new_val:
                            this_student_changes.append({
                                'type': f'调班-{label}',
                                'detail': f'{subject} {label}: {old_val} → {new_val}',
                                'field': field,
                                'old': old_val,
                                'new': new_val,
                                'subject': subject,
                            })

                    # 检测新增报科: 原台帐该科为 "-" 但现在是 "在读"
                    if old_status_norm == '-' and new_status == '在读':
                        this_student_changes.append({
                            'type': f'新增报科-{subject}',
                            'detail': f'新增{subject}在读（原无此科）',
                            'field': 'status',
                            'old': '-',
                            'new': '在读',
                            'subject': subject,
                        })

            # 是否续秋：秋季订单为真相源，无秋季订单时继承原台帐 T/AC/AL 列
            fall_entry = fall_data.get(uid, {}).get(subject, {})
            if isinstance(fall_entry, dict):
                fall_status = fall_entry.get('status', '')
                fall_pay_time = fall_entry.get('pay_time')
            else:
                # 向后兼容：旧格式为字符串
                fall_status = fall_entry
                fall_pay_time = None
            old_col = {'雪球思维': 'T', '悦读创作': 'AC', '双语素养': 'AL'}[subject]
            raw_cf = fall_status if fall_status else manual.get(old_col, '')
            new_cf = _fmt_continue_fall(raw_cf)
            student[f'{subject}_continue_fall'] = new_cf
            student[f'{subject}_fall_pay_time'] = fall_pay_time

            # 检测续秋变化（仅对非新增学员）
            if not is_new and uid in old_uids:
                old_cf_raw = manual.get(old_col, '')
                old_cf = _fmt_continue_fall(old_cf_raw)
                if old_cf != new_cf:
                    this_student_changes.append({
                        'type': f'续秋变化-{subject}',
                        'detail': f'{subject} 续秋: {old_cf} → {new_cf}',
                        'field': 'continue_fall',
                        'old': old_cf,
                        'new': new_cf,
                        'subject': subject,
                    })

            # 手动字段继承
            student[f'{subject}_score_source'] = manual.get('K', '') if subject == '雪球思维' else ''
            student[f'{subject}_pre_test'] = manual.get(
                {'雪球思维': 'L', '悦读创作': 'V', '双语素养': 'AE'}[subject], ''
            )
            student[f'{subject}_school_score'] = manual.get('M', '') if subject == '雪球思维' else ''

        # 整体状态
        status_parts = []
        if active_subjects:
            status_parts.append('在读')
        student['status'] = '、'.join(status_parts) if status_parts else '退费'

        sorted_periods = sorted(period_set, key=lambda p: PERIOD_ORDER.get(p, 99))
        student['period_combined'] = '/'.join(sorted_periods) if sorted_periods else '-'

        # 运营字段
        student['vip_group'] = manual.get('AM', '')
        student['visit_before_no'] = manual.get('AN', '')
        student['visit_before_confirm'] = manual.get('AO', '')
        student['visit_mid_no'] = manual.get('AP', '')
        student['visit_mid_confirm'] = manual.get('AQ', '')
        student['renew_willingness'] = manual.get('AR', '')
        student['review_basis'] = manual.get('AS', '')

        # 记录该学员的变动
        if is_new:
            enrolled_subjects = [s for s in ['雪球思维', '悦读创作', '双语素养']
                                if student.get(f'{s}_status', '-') not in ('-',)]
            refund_subjects = [s for s in ['雪球思维', '悦读创作', '双语素养']
                               if student.get(f'{s}_status', '-') in ('课前退', '课后退')]
            enrolled_str = '、'.join([f"{s}({student.get(f'{s}_class_id','-')})" for s in enrolled_subjects])
            refund_str = '、'.join([f"{s}({student.get(f'{s}_status','-')})" for s in refund_subjects])

            # 新增学员的详情：按学科展示期次/老师
            new_details = []
            for s in enrolled_subjects:
                new_details.append({
                    'type': f'新增报科-{s}',
                    'detail': f'{s} 在读（{student.get(f"{s}_class_id","-")}）',
                    'field': 'status',
                    'old': '-',
                    'new': '在读',
                    'subject': s,
                    'period': student.get(f'{s}_period_single', '-'),
                    'teacher': student.get(f'{s}_teacher', '-'),
                })
            for s in refund_subjects:
                new_details.append({
                    'type': f'新增退费-{s}',
                    'detail': f'{s} {student.get(f"{s}_status","-")}',
                    'field': 'status',
                    'old': '-',
                    'new': student.get(f'{s}_status', '-'),
                    'subject': s,
                    'period': student.get(f'{s}_period_single', '-'),
                    'teacher': student.get(f'{s}_teacher', '-'),
                })

            parts = []
            if enrolled_subjects:
                parts.append(f"报读: {enrolled_str}")
            if refund_subjects:
                parts.append(f"退费: {refund_str}")

            student_changes.append({
                'uid': uid,
                'name': first['name'],
                'teaching_point': first['teaching_point'],
                'change_type': '新增学员',
                'summary': '；'.join(parts) if parts else '新增学员（无报科）',
                'details': new_details,
                'is_new': True,
            })

            # 新增学员 → 逐条 change_details
            for subj in enrolled_subjects:
                change_details.append({
                    'uid': uid, 'name': first['name'], 'subject': subj,
                    'change_type': '新增学员',
                    'field_name': '全部',
                    'old_value': '-',
                    'new_value': f"{subj}-在读-{student.get(f'{subj}_class_id','')}",
                })
            for subj in refund_subjects:
                change_details.append({
                    'uid': uid, 'name': first['name'], 'subject': subj,
                    'change_type': '新增学员(退费)',
                    'field_name': '全部',
                    'old_value': '-',
                    'new_value': f"{subj}-{student.get(f'{subj}_status','')}",
                })

        elif this_student_changes:
            # 分组同类变动
            grouped = {}
            for ch in this_student_changes:
                key = ch['type']
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append(ch)

            summary_parts = []
            for ch_type, items in grouped.items():
                subj_list = '、'.join([it['subject'] for it in items])
                if '续秋变化' in ch_type:
                    # 续秋仅展示"续秋: 学科 旧→新"
                    for it in items:
                        summary_parts.append(f"续秋:{it['subject']} {it.get('old','-')}→{it.get('new','-')}")
                else:
                    summary_parts.append(f"{ch_type}({subj_list})")

            student_changes.append({
                'uid': uid,
                'name': first['name'],
                'teaching_point': first['teaching_point'],
                'change_type': '信息变更',
                'summary': '；'.join(summary_parts),
                'details': this_student_changes,
                'is_new': False,
            })

            # 逐条明细
            for ch in this_student_changes:
                change_details.append({
                    'uid': uid, 'name': first['name'], 'subject': ch.get('subject', ''),
                    'change_type': ch['type'],
                    'field_name': ch.get('field', ''),
                    'old_value': str(ch.get('old', '')),
                    'new_value': str(ch.get('new', '')),
                })

        students.append(student)

    # 排序：新学员置顶
    def sort_key(s):
        return (0 if s['is_new'] else 1,
                -(max((o['pay_time'] or datetime(2000, 1, 1)).timestamp()
                      for o in uid_orders.get(s['uid'], [])) or 0))

    students.sort(key=sort_key)

    # 统计各类变动
    new_count = sum(1 for s in students if s['is_new'])
    change_count = len(student_changes)
    # 按变动类型分组统计
    change_breakdown = defaultdict(int)
    for ch in change_details:
        change_breakdown[ch['change_type']] += 1

    stats = {
        'total_students': len(students),
        'active_students': sum(1 for s in students if '在读' in s.get('status', '')),
        'new_students': new_count,
        'new_refund_students': len(set(rd['uid'] for rd in new_refund_details)),
        'new_refund_details': len(new_refund_details),
        'all_refund_students': len(set(rd['uid'] for rd in all_refund_details)),
        'all_refund_details': len(all_refund_details),
        'changes_count': change_count,
        'change_breakdown': dict(change_breakdown),
        'calibration_date': datetime.now().strftime('%Y-%m-%d %H:%M'),
    }

    return {
        'students': students,
        'student_changes': student_changes,
        'change_details': change_details,
        'stats': stats,
        'refund_details': new_refund_details,
        'all_refund_details': all_refund_details,
        'anomalies': anomaly_records,
    }


def load_fall_orders(filepath):
    """加载秋季续费订单
    秋季订单文件列结构:
      col4=学生uid, col5=学生姓名, col7=班级名称, col24=学科,
      col31=在班状态, col32=出班原因, col36=支付时间, col38=是否有效退费
    返回: dict[uid][subject] = {'status': '是'/'退费'/'否', 'pay_time': datetime_or_None}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    fall = defaultdict(dict)
    for row in range(2, ws.max_row + 1):
        uid = str(ws.cell(row=row, column=4).value or '').strip()
        subject = str(ws.cell(row=row, column=24).value or '').strip()
        if not uid or not subject:
            continue
        # 标准化学科名称
        for std_subj in ('雪球思维', '悦读创作', '双语素养'):
            if std_subj in subject:
                subject = std_subj
                break
        else:
            continue  # 非三科学科，跳过

        status = str(ws.cell(row=row, column=31).value or '').strip()
        is_refund = str(ws.cell(row=row, column=38).value or '').strip()
        pay_time = parse_order_date(ws.cell(row=row, column=36).value)

        if status == '在读' and not is_refund:
            fall[uid][subject] = {'status': '是', 'pay_time': pay_time}
        elif '退费' in is_refund or '退费' in str(ws.cell(row=row, column=32).value or ''):
            fall[uid][subject] = {'status': '退费', 'pay_time': pay_time}
        else:
            fall[uid][subject] = {'status': '否', 'pay_time': pay_time}

    wb.close()
    return fall


# ============================================================
# Excel输出
# ============================================================

def generate_calibrated_ledger(calibration_result, template_file, output_path):
    """基于模板生成校准后的UV台帐"""
    import shutil
    shutil.copy2(template_file, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = _get_uv_sheet(wb)
    if ws is None:
        wb.close()
        return

    students = calibration_result['students']
    data_start = 3

    # 清除旧数据
    for row in range(data_start, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    font_normal = Font(name='微软雅黑', size=9)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, s in enumerate(students):
        row = data_start + i

        ws.cell(row=row, column=1, value=s['teaching_point'])
        ws.cell(row=row, column=2, value=s['uid'])
        ws.cell(row=row, column=3, value=s['name'])
        ws.cell(row=row, column=4, value=s.get('advisor', ''))
        ws.cell(row=row, column=5, value=s.get('channel', ''))
        ws.cell(row=row, column=6, value=s.get('purpose', ''))
        ws.cell(row=row, column=7, value=s.get('notes', ''))
        ws.cell(row=row, column=8, value=s['status'])
        ws.cell(row=row, column=9, value=s['period_combined'])

        # 雪球思维
        ws.cell(row=row, column=10, value=s.get('雪球思维_class_id', '-'))
        ws.cell(row=row, column=11, value=s.get('雪球思维_score_source', ''))
        ws.cell(row=row, column=12, value=s.get('雪球思维_pre_test', ''))
        ws.cell(row=row, column=13, value=s.get('雪球思维_school_score', ''))
        ws.cell(row=row, column=14, value=s.get('雪球思维_teacher', '-'))
        ws.cell(row=row, column=15, value=s.get('雪球思维_room', '-'))
        ws.cell(row=row, column=16, value=s.get('雪球思维_time_slot', '-'))
        ws.cell(row=row, column=17, value=s.get('雪球思维_class_type', '-'))
        ws.cell(row=row, column=18, value=s.get('雪球思维_period_single', '-'))
        ws.cell(row=row, column=19, value=s.get('雪球思维_status', '-'))
        ws.cell(row=row, column=20, value=s.get('雪球思维_continue_fall', ''))

        # 悦读创作
        ws.cell(row=row, column=21, value=s.get('悦读创作_class_id', '-'))
        ws.cell(row=row, column=22, value=s.get('悦读创作_pre_test', ''))
        ws.cell(row=row, column=23, value=s.get('悦读创作_teacher', '-'))
        ws.cell(row=row, column=24, value=s.get('悦读创作_room', '-'))
        ws.cell(row=row, column=25, value=s.get('悦读创作_time_slot', '-'))
        ws.cell(row=row, column=26, value=s.get('悦读创作_class_type', '-'))
        ws.cell(row=row, column=27, value=s.get('悦读创作_period_single', '-'))
        ws.cell(row=row, column=28, value=s.get('悦读创作_status', '-'))
        ws.cell(row=row, column=29, value=s.get('悦读创作_continue_fall', ''))

        # 双语素养
        ws.cell(row=row, column=30, value=s.get('双语素养_class_id', '-'))
        ws.cell(row=row, column=31, value=s.get('双语素养_pre_test', ''))
        ws.cell(row=row, column=32, value=s.get('双语素养_teacher', '-'))
        ws.cell(row=row, column=33, value=s.get('双语素养_room', '-'))
        ws.cell(row=row, column=34, value=s.get('双语素养_time_slot', '-'))
        ws.cell(row=row, column=35, value=s.get('双语素养_class_type', '-'))
        ws.cell(row=row, column=36, value=s.get('双语素养_period_single', '-'))
        ws.cell(row=row, column=37, value=s.get('双语素养_status', '-'))
        ws.cell(row=row, column=38, value=s.get('双语素养_continue_fall', ''))

        # 运营字段
        ws.cell(row=row, column=39, value=s.get('vip_group', ''))
        ws.cell(row=row, column=40, value=s.get('visit_before_no', ''))
        ws.cell(row=row, column=41, value=s.get('visit_before_confirm', ''))
        ws.cell(row=row, column=42, value=s.get('visit_mid_no', ''))
        ws.cell(row=row, column=43, value=s.get('visit_mid_confirm', ''))
        ws.cell(row=row, column=44, value=s.get('renew_willingness', ''))
        ws.cell(row=row, column=45, value=s.get('review_basis', ''))

        for col in range(1, 46):
            cell = ws.cell(row=row, column=col)
            cell.font = font_normal
            cell.alignment = align_center

        if s['is_new']:
            for col in range(1, 46):
                ws.cell(row=row, column=col).fill = green_fill

    wb.save(output_path)
    wb.close()


def generate_calibration_report(result, output_path):
    """
    生成台帐校准报告 (xlsx)
    Sheet: 校准概览 / 学员变动总览 / 变动明细 / 新增退费明细 / 重点关注
    """
    wb = openpyxl.Workbook()

    # 样式
    header_font = Font(name='微软雅黑', size=11, bold=True)
    section_font = Font(name='微软雅黑', size=10, bold=True, color='333333')
    normal_font = Font(name='微软雅黑', size=9)
    bold_font = Font(name='微软雅黑', size=9, bold=True)
    grey_fill = PatternFill(start_color='EDF0F5', end_color='EDF0F5', fill_type='solid')
    # 高饱和配色（直观清晰，2.0 风格）
    green_fill = PatternFill(start_color='C6E7CE', end_color='C6E7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FBE0BD', end_color='FBE0BD', fill_type='solid')
    red_fill = PatternFill(start_color='FAC9C9', end_color='FAC9C9', fill_type='solid')
    blue_fill = PatternFill(start_color='CFE0F7', end_color='CFE0F7', fill_type='solid')
    teal_fill = PatternFill(start_color='C9ECE6', end_color='C9ECE6', fill_type='solid')
    brand_fill = PatternFill(start_color='4F6BED', end_color='4F6BED', fill_type='solid')
    green_font = Font(name='微软雅黑', size=10, bold=True, color='1B7A3D')
    orange_font = Font(name='微软雅黑', size=10, bold=True, color='B5640A')
    red_font = Font(name='微软雅黑', size=10, bold=True, color='B7202E')
    blue_font = Font(name='微软雅黑', size=10, bold=True, color='234E9C')
    teal_font = Font(name='微软雅黑', size=10, bold=True, color='0B6B5E')
    brand_font = Font(name='微软雅黑', size=15, bold=True, color='FFFFFF')
    sub_font = Font(name='微软雅黑', size=9, color='64748B')
    sect_font = Font(name='微软雅黑', size=10, bold=True, color='475569')
    lbl_fill = PatternFill(start_color='F4F6FA', end_color='F4F6FA', fill_type='solid')
    lbl_font = Font(name='微软雅黑', size=9, color='475569')
    neu_fill = PatternFill(start_color='F7F9FC', end_color='F7F9FC', fill_type='solid')
    thin = Side(style='thin', color='D5DAE2')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cat_fill(label):
        s = str(label)
        if '新增学员' in s:
            return green_fill, green_font
        if '退费' in s:
            return red_fill, red_font
        if '调班' in s or '调期' in s:
            return orange_fill, orange_font
        if '续秋' in s or '续报' in s:
            return blue_fill, blue_font
        if '报科' in s:
            return teal_fill, teal_font
        return grey_fill, sect_font
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    stats = result['stats']
    new_refunds = result.get('refund_details', [])
    student_changes = result.get('student_changes', [])
    change_details = result.get('change_details', [])

    # ── Sheet 1: 校准概览 ──
    ws1 = wb.active
    ws1.title = '校准概览'
    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 16
    ws1.sheet_view.showGridLines = False

    # 标题带
    c = ws1.cell(row=1, column=1, value='UV台帳 校准报告')
    c.font = brand_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.merge_cells('A1:D1')
    ws1.row_dimensions[1].height = 32
    for col in range(1, 5):
        ws1.cell(row=1, column=col).fill = brand_fill
    c2 = ws1.cell(row=2, column=1, value=f"校准日期: {stats['calibration_date']}    暑 / 初一 / 三科")
    c2.font = sub_font

    def section_header(title, row):
        sc = ws1.cell(row=row, column=1, value=title)
        sc.font = sect_font
        sc.fill = grey_fill
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        for col in range(1, 5):
            ws1.cell(row=row, column=col).fill = grey_fill
        return row + 1

    r = 4
    # 核心指标
    r = section_header('核心指标', r)
    items = [
        ('学员总数', stats['total_students'], 'neu'),
        ('在读学员', stats['active_students'], 'green'),
        ('新增学员', stats['new_students'], 'green'),
        ('新增退费学员', stats['new_refund_students'], 'red'),
        ('新增退费单科数', stats['new_refund_details'], 'red'),
        ('存量退费学员', stats['all_refund_students'] - stats['new_refund_students'], 'neu'),
        ('存量退费单科数', stats['all_refund_details'] - stats['new_refund_details'], 'neu'),
        ('变动学员数', stats['changes_count'], 'orange'),
    ]
    for label, val, kind in items:
        lc = ws1.cell(row=r, column=1, value=label)
        lc.font = lbl_font
        lc.fill = lbl_fill
        lc.border = border
        lc.alignment = align_left
        vc = ws1.cell(row=r, column=2, value=val)
        if kind == 'green':
            f, ft = green_fill, green_font
        elif kind == 'red':
            f, ft = red_fill, red_font
        elif kind == 'orange':
            f, ft = orange_fill, orange_font
        else:
            f, ft = neu_fill, bold_font
        vc.fill = f
        vc.font = ft
        vc.border = border
        vc.alignment = align_center
        r += 1

    r += 1
    # 变动类型统计
    r = section_header('变动类型统计', r)
    h1 = ws1.cell(row=r, column=1, value='变动类型')
    h2 = ws1.cell(row=r, column=2, value='数量')
    for h in (h1, h2):
        h.font = bold_font
        h.fill = grey_fill
        h.border = border
        h.alignment = align_center
    r += 1
    breakdown = stats.get('change_breakdown', {})
    if breakdown:
        for ctype, cnt in sorted(breakdown.items(), key=lambda x: -x[1]):
            f, ft = cat_fill(ctype)
            a = ws1.cell(row=r, column=1, value=ctype)
            a.font = ft
            a.fill = f
            a.border = border
            a.alignment = align_left
            b = ws1.cell(row=r, column=2, value=cnt)
            b.font = bold_font
            b.fill = f
            b.border = border
            b.alignment = align_center
            r += 1
    else:
        a = ws1.cell(row=r, column=1, value='无变动')
        a.font = normal_font
        a.border = border
        r += 1

    r += 1
    # 退费比例
    r = section_header('退费比例（新增退费）', r)
    before_count = sum(1 for rd in new_refunds if rd.get('type') == '课前退')
    after_count = sum(1 for rd in new_refunds if rd.get('type') == '课后退')
    total_refund = before_count + after_count
    for label, val in [('新增课前退单科数', before_count), ('新增课后退单科数', after_count)]:
        lc = ws1.cell(row=r, column=1, value=label)
        lc.font = lbl_font
        lc.fill = lbl_fill
        lc.border = border
        lc.alignment = align_left
        vc = ws1.cell(row=r, column=2, value=val)
        vc.font = red_font
        vc.fill = red_fill
        vc.border = border
        vc.alignment = align_center
        r += 1
    if total_refund > 0:
        lc = ws1.cell(row=r, column=1, value='课前退占比')
        lc.font = lbl_font
        lc.fill = lbl_fill
        lc.border = border
        lc.alignment = align_left
        vc = ws1.cell(row=r, column=2, value=f'{before_count/total_refund*100:.0f}%')
        vc.font = red_font
        vc.fill = red_fill
        vc.border = border
        vc.alignment = align_center
        r += 1
    else:
        lc = ws1.cell(row=r, column=1, value='无新增退费')
        lc.font = normal_font
        lc.border = border
        r += 1

    r += 1
    # 重点关注
    r = section_header('重点关注', r)
    highlight_items = []
    for sc in student_changes:
        if sc['change_type'] == '新增学员':
            highlight_items.append((f"• 新增: {sc['name']}({sc['uid']}) {sc['teaching_point']} — {sc['summary']}", 'green'))
        elif '新增退费' in sc.get('summary', ''):
            highlight_items.append((f"• 退费: {sc['name']}({sc['uid']}) — {sc['summary']}", 'red'))
    if highlight_items:
        for h, kind in highlight_items[:20]:
            f, ft = (green_fill, green_font) if kind == 'green' else (red_fill, red_font)
            lc = ws1.cell(row=r, column=1, value=h)
            lc.font = normal_font
            lc.fill = f
            lc.border = border
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            for col in range(1, 5):
                ws1.cell(row=r, column=col).fill = f
            r += 1
    else:
        lc = ws1.cell(row=r, column=1, value='• 无异常变动')
        lc.font = normal_font
        r += 1

    r += 1
    legend = '图例： 新增 / 在读 = 绿    退费 = 红    调班 / 调期 = 橙    续秋 = 蓝    报科 = 青'
    lc = ws1.cell(row=r, column=1, value=legend)
    lc.font = Font(name='微软雅黑', size=8, color='94A3B8')
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    ws1.freeze_panes = 'A4'

    # ── Sheet 2: 学员变动总览 ──
    ws2 = wb.create_sheet('学员变动总览')
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 60

    ws2.cell(row=1, column=1, value='学员变动总览').font = Font(name='微软雅黑', size=12, bold=True)
    n_new = stats['new_students']
    ws2.cell(row=2, column=1,
             value=f"共 {stats['changes_count']} 名学员有变动 (新增{n_new}，变更{stats['changes_count']-n_new})"
             ).font = Font(name='微软雅黑', size=9, color='666666')

    headers2 = ['变动类型', '姓名', 'uid', '教学点', '二级分类', '变动详情']
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = grey_fill
        c.alignment = align_center

    row = 5
    for sc in student_changes:
        ct = sc['change_type']
        # 颜色标记（按变动类型/摘要归类，底色饱和度提高）
        row_fill, _ = cat_fill(ct + ' ' + sc.get('summary', ''))

        # 二级分类：提取 summary 中的主要类型
        summary = sc.get('summary', '')
        sub_type = summary.split('；')[0] if '；' in summary else summary

        ws2.cell(row=row, column=1, value=ct).font = normal_font
        ws2.cell(row=row, column=2, value=sc['name']).font = normal_font
        ws2.cell(row=row, column=3, value=sc['uid']).font = normal_font
        ws2.cell(row=row, column=4, value=sc.get('teaching_point', '')).font = normal_font
        ws2.cell(row=row, column=5, value=sub_type).font = normal_font
        ws2.cell(row=row, column=6, value=summary).font = normal_font

        for col in range(1, 7):
            ws2.cell(row=row, column=col).fill = row_fill
            ws2.cell(row=row, column=col).alignment = align_center if col <= 5 else align_left

        row += 1

    ws2.freeze_panes = 'A5'

    # ── Sheet 3: 变动明细 ──
    ws3 = wb.create_sheet('变动明细')
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 22
    ws3.column_dimensions['G'].width = 22

    ws3.cell(row=1, column=1, value='变动明细').font = Font(name='微软雅黑', size=12, bold=True)
    headers3 = ['学员姓名', 'uid', '学科', '变动类型', '字段', '原值', '新值']
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(row=3, column=i, value=h)
        c.font = header_font
        c.fill = grey_fill
        c.alignment = align_center

    for i, c in enumerate(change_details):
        row = 4 + i
        ws3.cell(row=row, column=1, value=c.get('name', '')).font = normal_font
        ws3.cell(row=row, column=2, value=c.get('uid', '')).font = normal_font
        ws3.cell(row=row, column=3, value=c.get('subject', '')).font = normal_font
        ws3.cell(row=row, column=4, value=c.get('change_type', '')).font = normal_font
        ws3.cell(row=row, column=5, value=c.get('field_name', '')).font = normal_font
        ws3.cell(row=row, column=6, value=str(c.get('old_value', ''))).font = normal_font
        ws3.cell(row=row, column=7, value=str(c.get('new_value', ''))).font = normal_font
        for col in range(1, 8):
            ws3.cell(row=row, column=col).alignment = align_center

    ws3.freeze_panes = 'A4'

    # ── Sheet 4: 新增退费明细 ──
    ws4 = wb.create_sheet('新增退费明细')
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 12
    ws4.column_dimensions['E'].width = 20
    ws4.column_dimensions['F'].width = 10
    ws4.column_dimensions['G'].width = 20

    ws4.cell(row=1, column=1, value='新增退费学员明细').font = Font(name='微软雅黑', size=12, bold=True)
    ws4.cell(row=2, column=1,
             value=f"共{stats['new_refund_students']}名学员, {stats['new_refund_details']}单科退费 "
                   f"(课前退{before_count}, 课后退{after_count})").font = Font(name='微软雅黑', size=9, color='666666')
    ws4.cell(row=3, column=1,
             value=f"存量退费: {stats['all_refund_students']-stats['new_refund_students']}名, "
                   f"{stats['all_refund_details']-stats['new_refund_details']}单科（未列入本表）"
             ).font = Font(name='微软雅黑', size=9, color='999999')

    headers4 = ['学生uid', '姓名', '学科', '退费类型', '退费时间', '期次', '备注']
    for i, h in enumerate(headers4, 1):
        c = ws4.cell(row=5, column=i, value=h)
        c.font = header_font
        c.fill = grey_fill
        c.alignment = align_center

    for i, rd in enumerate(new_refunds):
        row = 6 + i
        ws4.cell(row=row, column=1, value=rd.get('uid', '')).font = normal_font
        ws4.cell(row=row, column=2, value=rd.get('name', '')).font = normal_font
        ws4.cell(row=row, column=3, value=rd.get('subject', '')).font = normal_font
        ws4.cell(row=row, column=4, value=rd.get('type', '')).font = normal_font
        rt = rd.get('time')
        if rt and hasattr(rt, 'strftime'):
            ws4.cell(row=row, column=5, value=rt.strftime('%Y-%m-%d')).font = normal_font
        else:
            ws4.cell(row=row, column=5, value=str(rt or '')).font = normal_font
        ws4.cell(row=row, column=6, value=rd.get('period', '')).font = normal_font
        ws4.cell(row=row, column=7, value='').font = normal_font
        for col in range(1, 8):
            ws4.cell(row=row, column=col).alignment = align_center

    ws4.freeze_panes = 'A6'

    # ── Sheet 5: 重点关注 ──
    ws5 = wb.create_sheet('重点关注')
    ws5.column_dimensions['A'].width = 90
    ws5.cell(row=1, column=1, value='重点关注').font = Font(name='微软雅黑', size=12, bold=True)
    r = 3

    if student_changes:
        for sc in student_changes:
            if sc['change_type'] == '新增学员':
                detail = f"• 新增: {sc['name']}({sc['uid']}) {sc.get('teaching_point','')} — {sc['summary']}"
                ws5.cell(row=r, column=1, value=detail).font = normal_font
                r += 1
            elif '退费' in sc.get('summary', ''):
                detail = f"• 新增退费: {sc['name']}({sc['uid']}) — {sc['summary']}"
                ws5.cell(row=r, column=1, value=detail).font = normal_font
                r += 1
        # 调班调期
        for sc in student_changes:
            if '调班' in sc.get('summary', '') or '调期' in sc.get('summary', ''):
                detail = f"• 调班/调期: {sc['name']}({sc['uid']}) — {sc['summary']}"
                ws5.cell(row=r, column=1, value=detail).font = normal_font
                r += 1
    else:
        ws5.cell(row=r, column=1, value='• 本次校准无异常变动').font = normal_font

    wb.save(output_path)
    wb.close()


def generate_diff_report(prev_result, curr_result, output_path):
    """生成两次校准之间的差异报告"""
    prev_students = {s['uid']: s for s in prev_result['students']}
    curr_students = {s['uid']: s for s in curr_result['students']}

    new_uids = set(curr_students.keys()) - set(prev_students.keys())
    removed_uids = set(prev_students.keys()) - set(curr_students.keys())

    diffs = []
    for uid in set(curr_students.keys()) & set(prev_students.keys()):
        ps = prev_students[uid]
        cs = curr_students[uid]
        for subject in ['雪球思维', '悦读创作', '双语素养']:
            for field in ['class_id', 'teacher', 'room', 'time_slot', 'class_type', 'period_single']:
                key = f'{subject}_{field}'
                pv = ps.get(key, '')
                cv = cs.get(key, '')
                if str(pv) != str(cv):
                    diffs.append({
                        'uid': uid, 'name': cs['name'], 'subject': subject,
                        'field': field, 'old': pv, 'new': cv
                    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '差异对比'
    headers = ['uid', '姓名', '学科', '字段', '原值', '新值']
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)

    ws.cell(row=2, column=1, value=f"新增: {len(new_uids)} | 移除: {len(removed_uids)} | 变动: {len(diffs)}")

    for i, d in enumerate(diffs):
        row = 3 + i
        ws.cell(row=row, column=1, value=d['uid'])
        ws.cell(row=row, column=2, value=d['name'])
        ws.cell(row=row, column=3, value=d['subject'])
        ws.cell(row=row, column=4, value=d['field'])
        ws.cell(row=row, column=5, value=str(d['old']))
        ws.cell(row=row, column=6, value=str(d['new']))

    wb.save(output_path)
    wb.close()
    return {
        'new_uids': list(new_uids),
        'removed_uids': list(removed_uids),
        'diffs': diffs,
    }
