# -*- coding: utf-8 -*-
"""
订单明细表 → 学员快照 加载器（Phase 0–2 原型，只读 dry-run 安全）

设计要点：
- 按表头「名称」查找列（不依赖列序），对 uanme 笔误做归一化。
- 聚合粒度：原始行 = 学员×科目×订单；聚合为「学员行」(uid 为主键)，
  subjects_json 以 学科 为键（同科目多行合并：优先在读、期次并集）。
- 不修改生产库：本模块只解析/聚合/返回内存结构；dry-run 仅打印校验统计。
- 决策落地（见升级方案文档 §3/§7/§10）：
  * channel = 订单线索二级渠道
  * grade    = 班级年级
  * UID 跨分校唯一（合并层直接按 uid 聚合）
  * 其余 4 顾问角色留 enrich_json，不进筛选器
  * 期次"无" = 秋/春季学期 → effective_term 派生
"""
import os
import re
import json
import glob
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter

import openpyxl

# 合并物化依赖（repository 不反向 import 本模块，无循环依赖）
from repository import set_meta_value, get_meta_value, get_run_by_id, rebuild_active_run, _upgrade_old_subjects

from config import normalize_teaching_point
from engine.calibrator import normalize_branch

# ── 列头归一化：修正 uanme 笔误（商机创建人uanme 等）──
def _norm_header(h):
    if h is None:
        return ''
    h = str(h).strip()
    # 旧表存在 uanme 拼写错误，统一纠正为 uname（仅影响名称匹配，不影响取值）
    h = h.replace('uanme', 'uname')
    return h


# ── 关键字段 → 表头名称（用「纯名称」列，不用 uname 账号列）──
FIELD_COLS = {
    'uid':           '学生uid(归一化)',
    'name':          '学生姓名',
    'city_branch':   '城市分校',
    'teaching_point':'教学点',
    'room':          '教室',
    'year':          '学年',
    'season':        '学季',
    'period':        '期次',
    'grade':         '班级年级',
    'subject':       '学科',
    'teacher':       '班级主讲',
    'class_id':      '班级ID',
    'course_type':   '课程类型',
    'product_type':  '产品类型',
    'class_mode':    '课程模式',
    'class_type':    '班型',
    'time_slot':     'U2开课时间',
    'pay_time':      '支付时间',
    'refund_time':   '退费时间',
    'actual_income': '实际收入金额（实际支付金额-退费金额-退差价金额）',
    'enroll_status': '在班状态',
    'valid_refund':  '是否有效退费',
    'pre_refund':    '是否课前有效退费',
    'post_refund':   '是否课后有效退费',
    'refund_reason_cat': '退费原因分类',
    'refund_reason': '退费原因',
    'attendance':    '出勤讲次数（含补卡）',
    'advisor_deal_owner':   '成单归属人',
    'advisor_creator':      '商机创建人',
    'advisor_deal_follower':'成单商机跟进人',
    'advisor_goods_owner':  '订单商品码归属人',
    'advisor_course_follower':'课程跟进顾问',
    'channel_lead_l1': '订单线索一级渠道',
    'channel_lead_l2': '订单线索二级渠道',
    'channel_deal_l1': '订单商机一级渠道',
    'channel_deal_l2': '订单商机二级渠道',
    'order_origin':   '订单来源[新绩效类型](从暑开始算，仅对特惠和系统课判断)',
    'work_attr':      '做工归属',
    'entry_type':     '入班类型',
    'renew_next':     '是否续下一学季',
    'renew_next_type':'续下一学季类型',
    'retain_last':    '是否是上学季留存',
    'renew_denom_incl':  '是否应续分母（含联报）',
    'renew_denom_next':  '是否应续下学季分母（不含联报）',
    'renew_denom_next2': '是否应续下下学季分母（不含联报）',
    'school_grade':   '学校年级',
    'school_name':    '公立校名称',
    'uid_hash':       '学生hash手机号',
    'class_start':    '班级开始日期',
    'class_end':      '班级结束日期',
    'week_cycle':     '上课周期',
    'student_type_raw': '订单来源[新绩效类型](从暑开始算，仅对特惠和系统课判断)',  # BS列：拉新/老生/续报扩科
}


def _norm_branch(cb):
    """分校标准化：与 calibrator.normalize_branch 统一口径"""
    return normalize_branch(cb)


def _norm_tp(tp):
    """教学点标准化：与 config.normalize_teaching_point 保持一致。"""
    return normalize_teaching_point(tp)


# 学季规范化：新表值为 暑/寒/春/秋，前端筛选器选项为 暑假/寒假/春季/秋季
_SEASON_NORM = {'暑': '暑假', '寒': '寒假', '春': '春季', '秋': '秋季'}


def _norm_season(s):
    if not s:
        return ''
    s = str(s).strip()
    return _SEASON_NORM.get(s, s)


_SEASON_MAP = {'暑': '暑假', '寒': '寒假', '春': '春季', '秋': '秋季'}

# 期次排序：零<一<二<三<四<无（用于 period_combined 拼接）
_PERIOD_RANK = {'零期': 0, '一期': 1, '二期': 2, '三期': 3, '四期': 4, '无': 5}

# 学员归属映射（BS列「订单来源[新绩效类型]」→ UI 展示）
_STUDENT_TYPE_MAP = {
    '拉新': '新生',
    '老生': '老生',
    '续报扩科': '老生拓科',
}


def _map_student_type(raw):
    """BS列「订单来源[新绩效类型]」→ 学员归属（新生/老生/老生拓科）。
    取值集合：拉新/老生/续报扩科；其他值（空/-/未知）原样保留供排查。"""
    s = (raw or '').strip()
    if not s or s == '-':
        return ''
    return _STUDENT_TYPE_MAP.get(s, s)


def _norm_period(period):
    """期次规范化为底表 I列原始值（零期/一期/二期/三期/四期/无）。
    period 与 season 是两个独立维度：
      - period="无" 表示春/秋长学期（无期次之分），仍忠实保留为"无"。
      - 学季信息一律走独立的 season 字段，不再混入 period。
    这样期次筛选器可呈现 6 个基础标签，各维度可数据驱动联动。"""
    p = (period or '').strip()
    return p if p else '无'


def _clean(v):
    if v is None:
        return ''
    if isinstance(v, float) and v != v:  # NaN
        return ''
    s = str(v).strip()
    if s in ('-', 'None', 'nan', 'NaN'):
        return ''
    return s


def _is_yes(v):
    return _clean(v) in ('是', '1', 'true', 'True', 'Y', 'y')


def load_file(path):
    """读取订单明细表 → list[dict(字段名: 清洗值)]"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    raw_headers = [c.value for c in first_row]
    # 建立 归一化表头 → 列索引（同一归一名取首次出现）
    col_idx = {}
    for i, h in enumerate(raw_headers):
        nh = _norm_header(h)
        if nh and nh not in col_idx:
            col_idx[nh] = i  # 0-based
    # 反查 FIELD_COLS 中每个字段对应的列索引
    field_to_col = {}
    missing = []
    for fkey, hname in FIELD_COLS.items():
        nh = _norm_header(hname)
        if nh in col_idx:
            field_to_col[fkey] = col_idx[nh]
        else:
            missing.append(hname)
    if missing:
        raise ValueError(f"缺少必要列: {missing}")

    rows = []
    for r in ws.iter_rows(min_row=2):
        rec = {}
        for fkey, ci in field_to_col.items():
            rec[fkey] = _clean(r[ci].value)
        if not rec.get('uid'):
            continue  # 跳过无 uid 行
        rows.append(rec)
    wb.close()
    return rows


def _format_time_slot(v):
    """将 U2开课时间(datetime / 字符串)统一格式化为 HH:MM 上课时间。"""
    if v is None or v == '':
        return ''
    if isinstance(v, datetime):
        return v.strftime('%H:%M')
    s = str(v).strip()
    if not s or s in ('-', 'None', 'nan', 'NaN'):
        return ''
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M'):
        try:
            return datetime.strptime(s, fmt).strftime('%H:%M')
        except ValueError:
            continue
    return s


def _build_line(row):
    """将一条订单行(学员×科目×班级)转为一条独立的班级行(line)子结构。

    ⚠️ 粒度：不做任何跨行合并。每条订单行 = 一条 line，
    完整保留行级属性（year/season/grade/period/subject/status/...），
    使下游 PV = 满足条件的 line 数、UV = 去重 uid，完全对齐 Excel 逐行筛选。

    状态：直接采用「在班状态」原值（active 判定 = '在读' in status），
    与用户手动按 AQ列(在班状态)=在读 的口径一致。
    退费：valid_refund + refund_kind（单标签展示，课前优先）+ 独立标记
    pre_refund/post_refund（忠实映射 Excel 两列，可同时为是），均不覆盖 status。
    续报：'是否续下一学季'=是 → continue_fall='是'（以本行订单实际支付时间
    pay_time 作为 fall_pay_time，供下游按「支付时间 < 班级开始日期」判定
    开课前续报 / 当期转化）。
    """
    is_refund = _is_yes(row.get('valid_refund'))
    is_pre  = _is_yes(row.get('pre_refund'))
    is_post = _is_yes(row.get('post_refund'))
    # 单标签展示口径：课前退优先（一行只挂一个 refund_kind）
    if is_pre:
        refund_kind = '课前退'
    elif is_post:
        refund_kind = '课后退'
    elif is_refund:
        refund_kind = '退费'
    else:
        refund_kind = ''
    rn = (row.get('renew_next') or '').strip()
    line = {
        'subject':      row.get('subject', ''),
        'year':         row.get('year', ''),
        'season':       _norm_season(row.get('season', '')),
        'grade':        row.get('grade', ''),
        'period':       _norm_period(row.get('period')),   # 底表I列原值(零/一/二/三/四/无)
        'status':       row.get('enroll_status', ''),   # 在班状态原值
        'teacher':      row.get('teacher', ''),
        'class_id':     row.get('class_id', ''),
        'room':         row.get('room', ''),
        'teaching_point': _norm_tp(row.get('teaching_point', '')),
        'class_type':   row.get('class_type', ''),
        'course_type':  row.get('course_type', ''),
        'product_type': row.get('product_type', ''),
        'class_mode':   row.get('class_mode', ''),
        # ── 行级渠道（AL/AM 列）──
        # 每条订单行带自己的渠道，不能用学员级（学员级是从某行取的单一值，
        # 跨多渠道时会算错）。by_channel 累加时按行级 channel 取数。
        'channel_l1':   row.get('channel_lead_l1', ''),
        'channel_l2':   row.get('channel_lead_l2', ''),
        # 学员归属（BS 列「订单来源[新绩效类型]」）：
        #   拉新 → 新生 / 老生 → 老生（续报同科目）/ 续报扩科 → 老生拓科
        'student_type': _map_student_type(row.get('student_type_raw', '')),
        'time_slot':    _format_time_slot(row.get('time_slot', '')),   # U2开课时间 → HH:MM
        'week_cycle':   row.get('week_cycle', ''),
        'pay_time':     row.get('pay_time', ''),
        # 原生 AT 列「是否续下一学季」（后续会被跨学季判定改写 continue_fall，
        # 但本字段保留原值，供拉新×续报月度图直接取数）
        'renew_next':   rn,
        'continue_fall': '是' if rn == '是' else ('否' if rn == '否' else ''),
        'renew_next_type': row.get('renew_next_type', ''),
        'fall_pay_time': row.get('pay_time', ''),   # 本行订单实际支付时间，用于续报 pre/new 判定
        'valid_refund': '是' if is_refund else '',
        'refund_kind':  refund_kind,
        # 忠实映射 Excel 两个独立列（可同时为是；双标 39 行两维皆计入）
        'pre_refund':   '是' if is_pre else '',
        'post_refund':  '是' if is_post else '',
        'income':       _to_float(row.get('actual_income')),
        'attendance':   _to_int(row.get('attendance')),
        'start_date':   row.get('class_start', ''),
        'end_date':     row.get('class_end', ''),
        # 原生 BS 列「订单来源[新绩效类型]」逐行标签（拉新/续报扩科/老生）。
        # 拉新×续报月度图以该标签作为"纯拉新"判定依据（B 方案：以源表标签为准）。
        'order_origin': row.get('order_origin', ''),
        # BW 列「应续分母（含联报）」，续报率锁定分母口径使用
        'renew_denom_incl': row.get('renew_denom_incl', ''),
    }
    return line


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def aggregate(rows):
    """聚合为学员行（订单行数组粒度）。返回 dict[uid] = record。

    ⚠️ 核心：subjects_json 不再按学科去重，而是每条订单行(班级行)一个元素，
    以合成唯一键存放（key 无业务含义，真实学科在 line['subject']），
    每条 line 自带 year/season/grade/period/status 等行级属性。
    下游 PV = 满足条件的 line 数、UV = 去重 uid，完全对齐 Excel 逐行筛选。

    record = {
      uid, name, branch, teaching_point(主), grade(主/legacy), season(主/legacy),
      advisor, channel, status(汇总), period_combined,
      subjects_json: {synthetic_key: line}, enrich_json: {...}
    }
    """
    students = {}
    for row in rows:
        uid = row['uid']
        st = students.get(uid)
        if st is None:
            st = {
                'uid': uid,
                'name': row.get('name', ''),
                'branch': _norm_branch(row.get('city_branch')),
                'tp_counter': Counter(),
                'grade_counter': Counter(),
                'season_counter': Counter(),
                'advisor': row.get('advisor_deal_owner', ''),
                'channel': row.get('channel_lead_l2', ''),
                'lines': [],
                'rows': [],
            }
            students[uid] = st
        st['tp_counter'][row.get('teaching_point', '')] += 1
        st['grade_counter'][row.get('grade', '')] += 1
        st['season_counter'][_norm_season(row.get('season', ''))] += 1
        st['rows'].append(row)
        subj = row.get('subject', '')
        if not subj:
            continue
        st['lines'].append(_build_line(row))

    # 后处理：汇总顶层字段
    records = {}
    collision = 0  # 保留返回签名（行数组模型下无「同科目碰撞」概念）
    for uid, st in students.items():
        subjects_json = {}
        periods_all = set()
        any_reading = False
        any_valid_refund = False
        for i, line in enumerate(st['lines']):
            # 合成唯一键：学科#序号，保证不碰撞（真实学科在 line['subject']）
            key = f"{line['subject']}#{i}"
            p = line.get('period', '')
            if p:
                periods_all.add(p)
            if '在读' in (line.get('status') or ''):
                any_reading = True
            if line.get('valid_refund') == '是':
                any_valid_refund = True
            subjects_json[key] = line
        # 主教学点/年级/学季 = 众数（仅供展示 & legacy fallback，非筛选真相源）
        teaching_point = (st['tp_counter'].most_common(1)[0][0] if st['tp_counter'] else '')
        grade = (st['grade_counter'].most_common(1)[0][0] if st['grade_counter'] else '')
        season = (st['season_counter'].most_common(1)[0][0] if st['season_counter'] else '')
        # 汇总 status
        if any_reading:
            status = '在读'
        elif any_valid_refund:
            status = '有效退费'
        else:
            status = '出班'
        # enrich_json
        r0 = st['rows'][0]
        enrich = {
            'advisor_roles': {
                'creator': r0.get('advisor_creator', ''),
                'deal_follower': r0.get('advisor_deal_follower', ''),
                'goods_owner': r0.get('advisor_goods_owner', ''),
                'deal_owner': r0.get('advisor_deal_owner', ''),
                'course_follower': r0.get('advisor_course_follower', ''),
            },
            'channel_lead_l1': r0.get('channel_lead_l1', ''),
            'channel_lead_l2': r0.get('channel_lead_l2', ''),
            'channel_deal_l1': r0.get('channel_deal_l1', ''),
            'channel_deal_l2': r0.get('channel_deal_l2', ''),
            'income': sum(_to_float(x.get('actual_income', '')) for x in st['rows']),
            'order_origin': r0.get('order_origin', ''),
            'work_attr': r0.get('work_attr', ''),
            'entry_type': r0.get('entry_type', ''),
            'renew_next': r0.get('renew_next', ''),
            'renew_next_type': r0.get('renew_next_type', ''),
            'retain_last': r0.get('retain_last', ''),
            'renew_denom_incl': r0.get('renew_denom_incl', ''),
            'renew_denom_next': r0.get('renew_denom_next', ''),
            'renew_denom_next2': r0.get('renew_denom_next2', ''),
            'school_grade': r0.get('school_grade', ''),
            'school_name': r0.get('school_name', ''),
            'uid_hash': r0.get('uid_hash', ''),
            'refund_reason_cat': r0.get('refund_reason_cat', ''),
            'refund_reason': r0.get('refund_reason', ''),
        }
        records[uid] = {
            'uid': uid,
            'name': st['name'],
            'branch': st['branch'],
            'teaching_point': _norm_tp(teaching_point),
            'grade': grade,
            'season': season,
            'advisor': st['advisor'],
            'channel': st['channel'],
            'status': status,
            'period_combined': '/'.join(sorted(periods_all, key=lambda x: (_PERIOD_RANK.get(x, 99), x))),
            'subjects_json': subjects_json,
            'enrich_json': enrich,
        }
    return records, collision


# ═══════════════════════════════════════════════════════════════
# 跨学季续报判定（核心口径修正）
#   - 续报人员 = 当前学季在读 且 下个学季也有在读（同 uid 跨学季重叠）
#   - 某学科行(学季S, 学科N) 续报 ⇔ 该 uid 在下个学季 next_season(S) 有同名学科N 订单
#   - 续报行的「开课前续报 / 当期转化」用 下学季同名订单的最早支付时间
#     对比 当前学季开课日(该学季全部行最早 class_start) 判定：
#       支付时间 < 当前学季开课日 → 'pre' (开课前续报)
#       支付时间 >= 当前学季开课日 → 'new' (当期转化)
#   改写每行 line 的 continue_fall / fall_pay_time / start_date，
#   下游 engine/repo/api/exporter 零改动即自动正确。
# 学季循环：寒 → 春 → 暑 → 秋 → (次年)寒
# ═══════════════════════════════════════════════════════════════
SEASON_CYCLE = ['寒假', '春季', '暑假', '秋季']


def _next_season(s):
    try:
        i = SEASON_CYCLE.index(s)
    except ValueError:
        return None
    return SEASON_CYCLE[(i + 1) % len(SEASON_CYCLE)]


def _next_year_season(year, season):
    """学年循环中的下一个 (year, season)，按年份精确向后配对。

    寒假Y → 春季Y
    春季Y → 暑假Y
    暑假Y → 秋季Y
    秋季Y → 寒假(Y+1)

    关键修正：秋季Y 的下一季是次年寒假(Y+1)，而绝不能是同年的寒假Y
    （同年的寒假Y 在日历上早于秋季Y，是“上一季”，接回它属于时间倒挂）。
    """
    try:
        i = SEASON_CYCLE.index(season)
    except ValueError:
        return (year, season)
    if i == len(SEASON_CYCLE) - 1:  # 秋季 → 次年寒假
        ny = int(year) + 1 if str(year).isdigit() else year
        return (ny, SEASON_CYCLE[0])
    return (year, SEASON_CYCLE[i + 1])


def enrich_cross_season_renewal(records):
    """对 records(dict uid->record, 含 subjects_json) 做跨学季续报判定并就地改写。

    返回被判定为续报(跨学季重叠)的 (uid, 学科行key) 数量统计。

    配对口径（已修正年份倒挂）:
      以 (year, season) 为单位向后配对 —— 当前行(年Y, 学季S) 仅当其学员在
      _next_year_season(Y, S) 同名下有在读订单时才算续报人员。这样秋季Y 只会去
      匹配 寒假(Y+1)（未发生则不续报），而不会误接回早已过去的 寒假Y。
    """
    def _line_active(line):
        return '在读' in str(line.get('status', ''))

    # 1) 建立 uid -> {(year, season) -> {学科 -> [(key, line)]}}
    uid_idx = {}
    for uid, rec in records.items():
        by_ys = defaultdict(lambda: defaultdict(list))
        for key, line in (rec.get('subjects_json') or {}).items():
            y = line.get('year') or ''
            s = line.get('season') or ''
            n = line.get('subject') or ''
            by_ys[(y, s)][n].append((key, line))
        uid_idx[uid] = by_ys

    # 2) 每个 (year, season) 开课日 = 该学季所有「在读」行最早 class_start
    #    非在读行不参与续报判定，避免退费/出班数据把续报率撑高。
    season_start = {}
    for by_ys in uid_idx.values():
        for (y, s), by_name in by_ys.items():
            if (y, s) not in season_start:
                season_start[(y, s)] = None
            for n, items in by_name.items():
                for _key, line in items:
                    if not _line_active(line):
                        continue
                    sd = line.get('start_date') or line.get('class_start') or ''
                    if not sd:
                        continue
                    try:
                        d = datetime.fromisoformat(str(sd)).date()
                    except ValueError:
                        continue
                    if season_start[(y, s)] is None or d < season_start[(y, s)]:
                        season_start[(y, s)] = d

    # 3) 逐行判定并改写：只有「当前行在读」且「下学季同名学科存在在读行」才算续报。
    renewed_lines = 0
    for uid, by_ys in uid_idx.items():
        for (y, s), by_name in by_ys.items():
            ny, ns = _next_year_season(y, s)
            ns_by_name = by_ys.get((ny, ns), {})
            ref = season_start.get((y, s))
            for n, items in by_name.items():
                # 下学季同名学科仅取「在读」行
                ns_items = [it for it in ns_by_name.get(n, []) if _line_active(it[1])]
                if ns_items and ref is not None:
                    # 跨学季重叠 => 续报（无论下学季订单是否有支付时间，重叠即续报人员）
                    # 取下学季同名学科最早支付时间用于 pre/new 分类
                    npts = []
                    for _k, l in ns_items:
                        pt = l.get('pay_time') or l.get('fall_pay_time')
                        if pt:
                            try:
                                npts.append(datetime.fromisoformat(str(pt)))
                            except ValueError:
                                pass
                    if npts:
                        earliest = min(npts)
                        cls = 'new' if earliest.date() >= ref else 'pre'
                        renew_pay = earliest.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        # 无支付时间：保守归为开课前续报，fall_pay_time 置为开课日前一天以保证 classify→pre
                        cls = 'pre'
                        renew_pay = (ref - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
                    for _key, line in items:
                        line['continue_fall'] = '是'
                        line['fall_pay_time'] = renew_pay
                        line['start_date'] = ref.strftime('%Y-%m-%d')
                        line['renewal_class'] = cls
                        renewed_lines += 1
                else:
                    for _key, line in items:
                        line['continue_fall'] = '否'
    return renewed_lines


def find_order_detail_files(downloads_dir='/Users/wangboning/Downloads'):
    return sorted(glob.glob(os.path.join(downloads_dir, '订单明细*.xlsx')))


def main():
    import sys
    files = sys.argv[1:] or find_order_detail_files()
    if not files:
        print("未找到订单明细*.xlsx")
        return
    all_rows = []
    for f in files:
        rows = load_file(f)
        print(f"[读取] {os.path.basename(f)} → {len(rows)} 行")
        all_rows.extend(rows)
    print(f"[合计] 原始订单行 = {len(all_rows)}")
    records, collision = aggregate(all_rows)
    print(f"[聚合] 学员数(uid) = {len(records)}  同科目多行碰撞 = {collision}")

    # 校验统计
    branch_c = Counter(r['branch'] for r in records.values())
    tp_c = Counter(r['teaching_point'] for r in records.values())
    season_c = Counter(r['season'] for r in records.values())
    grade_c = Counter(r['grade'] for r in records.values())
    advisor_c = Counter(r['advisor'] for r in records.values())
    channel_c = Counter(r['channel'] for r in records.values())
    status_c = Counter(r['status'] for r in records.values())

    # PV = 学科行数；UV = 学员数；退费UV = 有效退费学员
    pv = sum(len(r['subjects_json']) for r in records.values())
    refund_uv = sum(1 for r in records.values()
                    if any(s.get('valid_refund') == '是' for s in r['subjects_json'].values()))
    renew_uv = sum(1 for r in records.values()
                   if r['enrich_json']['renew_next'] == '是')

    print("\n=== 学员级分布 ===")
    print("分校:", dict(branch_c))
    print("教学点:", dict(tp_c))
    print("学季:", dict(season_c))
    print("班级年级(top5):", grade_c.most_common(5))
    print("汇总状态:", dict(status_c))
    print(f"UID={len(records)}  PV(学科数)={pv}  退费UV={refund_uv}  续下学季UV={renew_uv}")
    print(f"顾问数={len(advisor_c)}  渠道(线索二级)数={len(channel_c)}")

    # 学科分布
    subj_c = Counter()
    for r in records.values():
        for s in r['subjects_json'].values():
            subj_c[s['subject']] += 1
    print("学科 PV 分布:", dict(subj_c))

    # 与表内「当学季学科数(未退费)」交叉校验（取首位行）
    # 抽样打印 3 个学员
    print("\n=== 抽样学员 ===")
    for uid, r in list(records.items())[:3]:
        print(f"  {r['name']}({uid}) [{r['branch']}/{r['teaching_point']}/{r['grade']}] "
              f"status={r['status']} period={r['period_combined']} advisor={r['advisor']}")
        for s in r['subjects_json'].values():
            print(f"     - {s['subject']}: {s.get('status')} teacher={s.get('teacher')} "
                  f"period={s.get('period')} refund={s.get('refund_kind','-')}")


def write_records(records, db_path, run_id):
    """将聚合结果写入 SQLite（生产表结构扩展版）。

    仅写入指定 db_path（测试用独立库，绝不触碰生产 uv_dashboard.db）。
    扩展列：enrich_json(学员级全量) + school_name/order_origin/work_attr/renew_denom/refund_flag
    （课程类型/产品类型/班型 随科目变化，保留在 subjects_json，查询用 json_extract）。
    """
    import sqlite3
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS student_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT,
            uid TEXT, name TEXT, teaching_point TEXT, advisor TEXT, channel TEXT,
            status TEXT, period_combined TEXT, subjects_json TEXT, manual_json TEXT,
            branch TEXT, grade TEXT, season TEXT,
            enrich_json TEXT,
            school_name TEXT, order_origin TEXT, work_attr TEXT,
            renew_denom TEXT, refund_flag TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ss_run ON student_snapshots(run_id)")
    conn.execute("DELETE FROM student_snapshots WHERE run_id=?", (run_id,))
    for uid, r in records.items():
        en = r['enrich_json']
        has_refund = any(s.get('valid_refund') == '是' for s in r['subjects_json'].values())
        refund_flag = '有效退费' if has_refund else ''
        conn.execute("""
            INSERT INTO student_snapshots
              (run_id,uid,name,teaching_point,advisor,channel,status,period_combined,
               subjects_json,manual_json,branch,grade,season,enrich_json,
               school_name,order_origin,work_attr,renew_denom,refund_flag)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            run_id, r['uid'], r['name'], r['teaching_point'], r['advisor'], r['channel'],
            r['status'], r['period_combined'],
            json.dumps(r['subjects_json'], ensure_ascii=False), '{}',
            r['branch'], r['grade'], r['season'],
            json.dumps(en, ensure_ascii=False),
            en.get('school_name', ''), en.get('order_origin', ''), en.get('work_attr', ''),
            en.get('renew_denom_incl', ''), refund_flag,
        ))
    conn.commit()
    n = conn.execute("SELECT COUNT(*) FROM student_snapshots WHERE run_id=?", (run_id,)).fetchone()[0]
    conn.close()
    return n


def verify_db(db_path, run_id):
    """从测试库回读，校验写入一致性。"""
    import sqlite3
    conn = sqlite3.connect(db_path)
    rows = conn.execute("SELECT uid,subjects_json,enrich_json,branch,teaching_point,advisor,channel,status FROM student_snapshots WHERE run_id=?", (run_id,)).fetchall()
    conn.close()
    n_stu = len(rows)
    n_subj = 0
    branches = set()
    for r in rows:
        branches.add(r[3])
        n_subj += len(json.loads(r[1]))
    return {'students': n_stu, 'subjects': n_subj, 'branches': branches}


def import_to_production(db_path, files, run_id=None, set_current=True):
    """合并多分校订单明细表 → 写入生产库为新 run，并可选 pin 为当前基线。

    只 ADD 数据，不删除任何历史 run / 快照；可逆（清除 current_run_id pin 即回退）。
    - 自动 ALTER 补齐扩展列（enrich_json/school_name/order_origin/work_attr/renew_denom/refund_flag）。
    - Plan A 聚合：按 (uid,学科) 合并多班。
    - 分校增量合并：保留旧 base run 中非本次分校的学员，避免新分校数据覆盖旧分校。
    """
    import sqlite3
    if isinstance(files, str):
        files = [files]
    all_rows = []
    for f in files:
        all_rows.extend(load_file(f))
    records, collision = aggregate(all_rows)
    # 跨学季续报判定（改写 continue_fall/fall_pay_time/start_date）
    renewed = enrich_cross_season_renewal(records)
    if run_id is None:
        run_id = 'od_' + datetime.now().strftime('%Y%m%d_%H%M%S')

    # ── 分校增量合并：保留旧 base run 中非本次分校的学员 ──
    new_branches = {r['branch'] for r in records.values() if r.get('branch')}
    old_base_run_id = get_meta_value('base_run_id')
    kept_from_old = 0
    if old_base_run_id and new_branches:
        old_run = get_run_by_id(old_base_run_id)
        if old_run and old_run.get('students'):
            for s in old_run['students']:
                sb = s.get('branch', '')
                # 对旧数据也做一次标准化（兼容历史遗留的'K9广州分校'等写法）
                sb = normalize_branch(sb) if sb else ''
                # 保留不属于本次上传分校的学员（新分校数据优先，同 uid 会被 records 覆盖）
                if sb and sb not in new_branches and s.get('uid') not in records:
                    # 将 DB 格式转为 records 格式（subjects_json / enrich_json 从字符串→dict）
                    sj = s.get('subjects_json', '{}')
                    ej = s.get('enrich_json', '{}')
                    try:
                        sj = json.loads(sj) if isinstance(sj, str) else sj
                    except (json.JSONDecodeError, TypeError):
                        sj = {}
                    try:
                        ej = json.loads(ej) if isinstance(ej, str) else ej
                    except (json.JSONDecodeError, TypeError):
                        ej = {}
                    # 注入学员级 channel_l1（DB 列里没有，从 enrich_json 取）
                    if ej and not s.get('channel_l1'):
                        s['channel_l1'] = ej.get('channel_lead_l1', '') or ''
                    # 旧格式兼容：若 subjects_json 键不含 #suffix 则升级
                    sj = _upgrade_old_subjects(sj, s)
                    records[s['uid']] = {
                        'uid': s['uid'],
                        'name': s.get('name', ''),
                        'branch': sb,
                        'teaching_point': s.get('teaching_point', ''),
                        'advisor': s.get('advisor', ''),
                        'channel': s.get('channel', ''),
                        'grade': s.get('grade', ''),
                        'season': s.get('season', ''),
                        'status': s.get('status', ''),
                        'period_combined': s.get('period_combined', ''),
                        'subjects_json': sj,
                        'enrich_json': ej,
                    }
                    kept_from_old += 1

    # 1) 补齐扩展列（生产库可能缺）
    conn = sqlite3.connect(db_path)
    existing_cols = [r[1] for r in conn.execute(
        "PRAGMA table_info(student_snapshots)").fetchall()]
    for add in ('enrich_json', 'school_name', 'order_origin', 'work_attr',
                'renew_denom', 'refund_flag'):
        if add not in existing_cols:
            conn.execute(f"ALTER TABLE student_snapshots ADD COLUMN {add} TEXT")
    conn.commit()
    conn.close()

    # 2) 写学员快照（独立连接，已 commit 的 schema 变更对后续连接可见）
    n = write_records(records, db_path, run_id)

    # 3) 写 calibration_runs（使新 run 成为 2.0 默认最新；pin 后覆盖）
    active = sum(1 for r in records.values() if r['status'] == '在读')
    refund = sum(1 for r in records.values()
                 if any(s.get('valid_refund') == '是' for s in r['subjects_json'].values()))
    bc = Counter(r['branch'] for r in records.values())
    sc = Counter()
    for r in records.values():
        for s in r['subjects_json'].values():
            sc[s['subject']] += 1
    stats = {'total': len(records), 'active': active, 'refund': refund,
             'collision': collision, 'branch': dict(bc), 'subject': dict(sc),
             'kept_from_old_base': kept_from_old}
    run_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(db_path)
    conn.execute("""
        INSERT INTO calibration_runs
          (run_id, run_time, total_students, active_students, refund_students,
           new_students, changes_count, stats_json, source)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (run_id, run_time, len(records), active, refund, 0, 0,
          json.dumps(stats, ensure_ascii=False), 'full'))
    conn.commit()
    conn.close()

    # 4) 单管道模式：全量上传为唯一数据源，直接 pin 本 run。
    #    不再调用 rebuild_active_run（校准已降级为纯校准，不写 overlay）。
    #    分校隔离：本次上传分校的数据完全替换，其他分校从旧 base 保留（kept_from_old_base）。
    if set_current:
        from repository import set_current_run_id
        set_meta_value('base_run_id', run_id)
        set_current_run_id(run_id)
    return {'run_id': run_id, 'students': n, 'active': active, 'refund': refund,
            'stats': stats, 'rebuild': None, 'kept_from_old_base': kept_from_old}



if __name__ == '__main__':
    main()
