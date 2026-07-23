"""
UV Dashboard 2.0 — Excel 导出模块
将看板数据导出为带样式的 Excel（openpyxl）。
三种类型：students / matrix / schedule
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from repository import (
    get_filtered_students, get_period_point_matrix, get_class_schedule,
)
from config import (
    SUBJECTS, SUBJECT_LABEL, PERIOD_ORDER,
    is_continue_fall, classify_renewal_by_pay_time, get_line_subject,
    is_renewed_by_source,
)

# ── 样式 ──
_HEADER_FILL = PatternFill('solid', fgColor='4F6BED')
_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_TOTAL_FILL = PatternFill('solid', fgColor='EEF1FB')
_TOTAL_FONT = Font(bold=True, color='1F2937', size=11)
_THIN = Side(style='thin', color='D8DCE6')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

_SUBJ_LABELS = [SUBJECT_LABEL.get(s, s) for s in SUBJECTS]
_SUBJ_HEADER = '学科班量(' + '/'.join(_SUBJ_LABELS) + ')'


def _ts() -> str:
    d = datetime.now()
    return f'{d.year}{d.month:02d}{d.day:02d}_{d.hour:02d}{d.minute:02d}'


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _style_total_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.alignment = _CENTER if c > 1 else _LEFT
        cell.border = _BORDER


def _style_data_borders(ws, first_data_row, ncols):
    for r in range(first_data_row, ws.max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            if cell.alignment.horizontal is None:
                cell.alignment = _LEFT if c == 2 else _CENTER


def _autosize(ws, max_w=42):
    for col in ws.columns:
        length = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                s = str(cell.value)
                ln = sum(2 if ord(ch) > 127 else 1 for ch in s)
                length = max(length, ln)
        ws.column_dimensions[letter].width = min(max(length + 2, 8), max_w)


# ═══════════════════════════════════════════════════
# 学员明细导出
# ═══════════════════════════════════════════════════

def _renewal_label(sd: dict, period: str = '', start_date: str = '') -> str:
    """续报分类标签——锁定分母口径：用 is_renewed_by_source 判定是否续报，
    分类(pre/new)仍用 is_continue_fall + pay_time。"""
    if not is_renewed_by_source(sd):
        return '未续报'
    cf = sd.get('continue_fall', '')
    first_p = period.split('/')[0].strip() if period else ''
    cls = classify_renewal_by_pay_time(cf, first_p, start_date=start_date) if is_continue_fall(cf) else ''
    if cls == 'pre':
        return '开课前续报'
    if cls == 'new':
        return '当期转化'
    return '已续报'


def _build_students_wb(students):
    wb = Workbook()
    ws = wb.active
    ws.title = '学员明细'
    headers = ['UID', '姓名', '教学点', '顾问', '渠道', '学科', '班级ID', '主讲',
               '期次', '班型', '时段', '教室', '在读状态', '续秋', '续报分类', '支付时间']
    ws.append(headers)
    _style_header(ws, len(headers))

    for s in students:
        subs = s.get('subjects', {}) or {}
        if not subs:
            ws.append([s.get('uid', ''), s.get('name', ''), s.get('teaching_point', ''),
                       s.get('advisor', ''), s.get('channel', ''), '', '', '', '', '', '', '',
                       s.get('status', ''), '', '', ''])
            continue
        for subj_key, sd in subs.items():
            if not isinstance(sd, dict):
                continue
            real_subj = get_line_subject(sd, subj_key)
            label = SUBJECT_LABEL.get(real_subj, real_subj)
            cf = sd.get('continue_fall', '')
            period = (sd.get('period') or '').strip()
            ws.append([
                s.get('uid', ''), s.get('name', ''), s.get('teaching_point', ''),
                s.get('advisor', ''), s.get('channel', ''),
                label, sd.get('class_id', ''), sd.get('teacher', ''),
                period, sd.get('class_type', ''), sd.get('time_slot', ''), sd.get('room', ''),
                sd.get('status', ''),
                '是' if is_renewed_by_source(sd) else '否',
                _renewal_label(sd, period, sd.get('start_date', '')),
                sd.get('fall_pay_time', ''),
            ])

    _style_data_borders(ws, 2, len(headers))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'
    _autosize(ws)
    return _save(wb)


# ═══════════════════════════════════════════════════
# 期次×教学点矩阵导出
# ═══════════════════════════════════════════════════

def _build_matrix_wb(matrix_data):
    wb = Workbook()
    mtype = matrix_data.get('matrix_type', 'short_term')
    season = matrix_data.get('season_label', '')
    columns = matrix_data.get('columns', [])
    col_keys = [c['key'] for c in columns]
    m = matrix_data.get('matrix', {})
    points = matrix_data.get('points', [])

    ws = wb.active
    ws.title = f'{season}矩阵' if season else '期次矩阵'
    dim_label = '期次' if mtype == 'short_term' else '星期'
    headers = ['教学点', dim_label, '在读', '已续报', 'UV', '班级数', '续报率%', _SUBJ_HEADER]
    ws.append(headers)
    _style_header(ws, len(headers))

    def subj_str(d):
        return '/'.join(str(d.get(s, 0)) for s in SUBJECTS)

    total_rows = []
    for pt in points:
        for ck in col_keys:
            cell = m.get(pt, {}).get(ck, {})
            if not cell:
                continue
            ws.append([pt, ck, cell.get('active', 0), cell.get('renewed', 0),
                       cell.get('uv', 0), cell.get('class_count', 0),
                       cell.get('renewal_rate', 0), subj_str(cell.get('subject_classes', {}))])
        # 教学点合计
        row_uv = m.get('__row_uv__', {}).get(pt, 0)
        row_cls = m.get('__row_cls__', {}).get(pt, 0)
        tot_active = sum(m.get(pt, {}).get(ck, {}).get('active', 0) for ck in col_keys)
        tot_renewed = sum(m.get(pt, {}).get(ck, {}).get('renewed', 0) for ck in col_keys)
        tot_rate = round(tot_renewed / tot_active * 100, 1) if tot_active else 0
        ws.append([f'{pt} 合计', f'全部{dim_label}', tot_active, tot_renewed, row_uv, row_cls,
                   tot_rate, subj_str(m.get('__row_subj_cls__', {}).get(pt, {}))])
        total_rows.append(ws.max_row)

    # 列合计
    for ck in col_keys:
        col_uv = m.get('__col_uv__', {}).get(ck, 0)
        col_cls = m.get('__col_cls__', {}).get(ck, 0)
        tot_active = sum(m.get(pt, {}).get(ck, {}).get('active', 0) for pt in points)
        tot_renewed = sum(m.get(pt, {}).get(ck, {}).get('renewed', 0) for pt in points)
        tot_rate = round(tot_renewed / tot_active * 100, 1) if tot_active else 0
        ws.append(['全部教学点', ck, tot_active, tot_renewed, col_uv, col_cls,
                   tot_rate, subj_str(m.get('__col_subj_cls__', {}).get(ck, {}))])
        total_rows.append(ws.max_row)

    # 总计
    grand_uv = m.get('__grand_uv__', 0)
    grand_cls = m.get('__grand_cls__', 0)
    tot_active = sum(m.get(pt, {}).get(ck, {}).get('active', 0) for pt in points for ck in col_keys)
    tot_renewed = sum(m.get(pt, {}).get(ck, {}).get('renewed', 0) for pt in points for ck in col_keys)
    tot_rate = round(tot_renewed / tot_active * 100, 1) if tot_active else 0
    ws.append(['总计', f'全部{dim_label}', tot_active, tot_renewed, grand_uv, grand_cls,
               tot_rate, subj_str(m.get('__grand_subj_cls__', {}))])
    total_rows.append(ws.max_row)

    _style_data_borders(ws, 2, len(headers))
    for r in total_rows:
        _style_total_row(ws, r, len(headers))
    ws.freeze_panes = 'A2'
    _autosize(ws)
    return _save(wb)


# ═══════════════════════════════════════════════════
# 排班看板导出
# ═══════════════════════════════════════════════════

def _build_schedule_wb(schedule_data):
    wb = Workbook()
    ws = wb.active
    ws.title = '排班看板'
    headers = ['班级ID', '学科', '主讲', '期次', '时段', '班型', '教室', '教学点',
               '在读人数', '已续报', '续报率%', '学员']
    ws.append(headers)
    _style_header(ws, len(headers))

    for c in schedule_data.get('classes', []):
        names = '、'.join(st.get('name', '') for st in c.get('students', []))
        ws.append([
            c.get('class_id', ''), SUBJECT_LABEL.get(c.get('subject', ''), c.get('subject', '')),
            c.get('teacher', ''), c.get('period', ''), c.get('time_slot', ''),
            c.get('class_type', ''), c.get('room', ''), c.get('teaching_point', ''),
            c.get('total', 0), c.get('renewed_count', 0), c.get('renewal_rate', 0), names,
        ])

    _style_data_borders(ws, 2, len(headers))
    ws.freeze_panes = 'A2'
    _autosize(ws)
    return _save(wb)


# ═══════════════════════════════════════════════════
# 统一保存
# ═══════════════════════════════════════════════════

def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════
# 对外入口
# ═══════════════════════════════════════════════════

def export_excel(export_type: str, run_id: str, filters: dict = None,
                 filter_spec=None, schedule_params: dict = None):
    """导出指定类型的数据为 Excel，返回 (BytesIO, 文件名)"""
    if export_type == 'students':
        students = get_filtered_students(run_id, filters or {})
        return _build_students_wb(students), f'学员明细_{_ts()}.xlsx'
    elif export_type == 'matrix':
        data = get_period_point_matrix(run_id, filter_spec)
        return _build_matrix_wb(data), f'期次矩阵_{_ts()}.xlsx'
    elif export_type == 'schedule':
        data = get_class_schedule(run_id, **(schedule_params or {}))
        return _build_schedule_wb(data), f'排班看板_{_ts()}.xlsx'
    else:
        raise ValueError(f'Unknown export_type: {export_type}')
